"""历史送货：Excel 导入、分页、批量删除。"""

from __future__ import annotations

import io
import re
import zipfile
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation
from typing import Any, ClassVar

import pandas as pd
from sqlalchemy import and_, delete, desc, func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.intelligent_prediction.exceptions import ValidationBusinessException
from app.intelligent_prediction.logging_utils import get_logger
from app.intelligent_prediction.models import DeliveryRecord
from app.intelligent_prediction.services.weather_client import fetch_weather_for_delivery
from app.intelligent_prediction.schemas.history import (
    DeliveryRecordRead,
    DeliveryRecordUpdate,
    HistoryImportResponse,
    HistoryImportRowError,
    HistoryListResponse,
    HistoryQueryParams,
    HistoryStatsBucket,
    HistoryStatsResponse,
)

logger = get_logger(__name__)

# 送货日期：年-月-日 或 年/月/日（月日可一位数）；与 Excel 序列日互不冲突的数值范围
_DATE_YMD_SEP = re.compile(r"^(\d{4})[-/](\d{1,2})[-/](\d{1,2})(?:\s|$)")
# 中文：2026年1月9日；或仅 1月9日（缺省年份取 UTC 当日年份）
_DATE_CN_YMD = re.compile(r"^(\d{4})年(\d{1,2})月(\d{1,2})(?:日|号|號)?")
_DATE_CN_MD = re.compile(r"^(\d{1,2})月(\d{1,2})(?:日|号|號)?")
_EXCEL_SERIAL_MAX = 200_000  # 约到 2448 年，覆盖正常业务日期列


class HistoryService:
    """历史记录业务逻辑。"""

    # 必填列含「节假日」；选填「天气」默认晴；配置天气 API 后仍会写入 weather_json
    _REQUIRED_IMPORT_COLUMN_PAIRS: ClassVar[tuple[tuple[str, str], ...]] = (
        ("大区经理", "regional_manager"),
        ("冶炼厂", "smelter"),
        ("仓库", "warehouse"),
        ("送货日期", "delivery_date"),
        ("节假日", "holiday_yes_no"),
        ("品种", "product_variety"),
        ("重量", "weight"),
    )
    _OPTIONAL_IMPORT_COLUMN_PAIRS: ClassVar[tuple[tuple[str, str], ...]] = (
        ("仓库地址", "warehouse_address"),
        ("冶炼厂地址", "smelter_address"),
        ("天气", "import_weather"),
    )
    _ALL_IMPORT_COLUMN_PAIRS: ClassVar[tuple[tuple[str, str], ...]] = (
        _REQUIRED_IMPORT_COLUMN_PAIRS + _OPTIONAL_IMPORT_COLUMN_PAIRS
    )
    REQUIRED_COLUMNS_CANONICAL: dict[str, str] = dict(_REQUIRED_IMPORT_COLUMN_PAIRS)
    OPTIONAL_COLUMNS_CANONICAL: dict[str, str] = dict(_OPTIONAL_IMPORT_COLUMN_PAIRS)
    HEADER_TO_FIELD: ClassVar[dict[str, str]] = {
        **dict(_REQUIRED_IMPORT_COLUMN_PAIRS),
        **dict(_OPTIONAL_IMPORT_COLUMN_PAIRS),
    }
    _HEADER_TO_EXCEL_COLUMN: ClassVar[dict[str, str]] = {
        cn: chr(ord("A") + i) for i, (cn, _) in enumerate(_ALL_IMPORT_COLUMN_PAIRS)
    }

    @classmethod
    def import_template_headers(cls) -> list[str]:
        """与 GET /delivery-history/模板 生成的 xlsx/csv 表头顺序一致。"""
        return [cn for cn, _ in cls._ALL_IMPORT_COLUMN_PAIRS]

    ALIAS_TO_CANONICAL: dict[str, str] = {
        "大区经理": "大区经理",
        "大區經理": "大区经理",
        "Regional Manager": "大区经理",
        "regional_manager": "大区经理",
        "冶炼厂": "冶炼厂",
        "Smelter": "冶炼厂",
        "smelter": "冶炼厂",
        "仓库": "仓库",
        "倉庫": "仓库",
        "Warehouse": "仓库",
        "warehouse": "仓库",
        "送货日期": "送货日期",
        "送貨日期": "送货日期",
        "Delivery Date": "送货日期",
        "delivery_date": "送货日期",
        "到货日期": "送货日期",
        "到貨日期": "送货日期",
        "Arrival Date": "送货日期",
        "arrival_date": "送货日期",
        "节假日": "节假日",
        "節假日": "节假日",
        "Holiday": "节假日",
        "holiday": "节假日",
        "是否节假日": "节假日",
        "品种": "品种",
        "品種": "品种",
        "Product Variety": "品种",
        "product_variety": "品种",
        "重量": "重量",
        "Weight": "重量",
        "weight": "重量",
        "仓库地址": "仓库地址",
        "倉庫地址": "仓库地址",
        "Warehouse Address": "仓库地址",
        "warehouse_address": "仓库地址",
        "冶炼厂地址": "冶炼厂地址",
        "冶煉廠地址": "冶炼厂地址",
        "Smelter Address": "冶炼厂地址",
        "smelter_address": "冶炼厂地址",
        "天气": "天气",
        "天氣": "天气",
        "Weather": "天气",
        "weather": "天气",
    }

    #: 下载模板中示例行的大区经理须以此开头；导入时跳过，计入 skipped
    TEMPLATE_EXAMPLE_RM_PREFIX: ClassVar[str] = "(示例)"

    _TEMPLATE_HEADER_COMMENTS: ClassVar[dict[str, str]] = {
        "大区经理": (
            "必填。若值以「(示例)」开头，本行仅为模板示例，导入时自动跳过（不计错误）。"
        ),
        "冶炼厂": "选填；填写时长度不超过 100 字符。",
        "仓库": "必填。",
        "送货日期": (
            "必填。支持 YYYY-MM-DD、YYYY/M/D、YYYY年M月D日、M月D日（缺省年为当年 UTC）；"
            "亦支持 Excel 日期或序列号。"
        ),
        "节假日": "必填；仅可填「是」（当日为非工作日）或「否」（当日为工作日），须与业务一致。",
        "品种": "必填。",
        "重量": "必填；非负数字，可含小数。",
        "仓库地址": "选填；用于天气查询定位，建议填写完整地址，最长 512 字符。",
        "冶炼厂地址": "选填；用于天气查询定位，最长 512 字符。",
        "天气": "选填；不填或空则按「晴」；最长 64 字符。",
    }

    @staticmethod
    def import_template_xlsx_bytes() -> bytes:
        """生成含「导入数据」示例行与「使用说明」的 xlsx 模板字节流。"""
        from openpyxl import Workbook
        from openpyxl.comments import Comment
        from openpyxl.styles import Alignment, Font

        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "导入数据"
        cols = HistoryService.import_template_headers()
        ws.append(cols)
        for col_idx, name in enumerate(cols, start=1):
            cell = ws.cell(row=1, column=col_idx)
            hint = HistoryService._TEMPLATE_HEADER_COMMENTS.get(name)
            if hint:
                cell.comment = Comment(hint, "PD")
            cell.font = Font(bold=True)
        example_rows: list[tuple[Any, ...]] = [
            (
                "(示例)李经理",
                "金利",
                "华东一号仓",
                "2026-01-15",
                "否",
                "阴极铜",
                12.5,
                "上海市浦东新区示例路1号",
                "江苏省苏州市示例冶炼厂路2号",
                "多云",
            ),
            ("(示例)李经理", "", "华东一号仓", "2026/1/16", "否", "A级电解铜", 8.0, "", "", ""),
            (
                "(示例)王经理",
                "某冶炼厂",
                "华南中心库",
                "2026-01-17",
                "是",
                "铝锭",
                3.25,
                "广州市南沙区示例仓3号",
                "",
                "晴",
            ),
        ]
        for r in example_rows:
            ws.append(list(r))
        ws.freeze_panes = "A2"
        for col_letter, width in (
            ("A", 22),
            ("B", 12),
            ("C", 16),
            ("D", 14),
            ("E", 10),
            ("F", 14),
            ("G", 10),
            ("H", 28),
            ("I", 28),
            ("J", 10),
        ):
            ws.column_dimensions[col_letter].width = width

        ws_help = wb.create_sheet("使用说明", 1)
        help_lines = (
            "送货历史导入 — 使用说明\n\n"
            "一、工作表\n"
            "·「导入数据」：系统仅读取此工作表（须为工作簿中的第一个工作表）。请勿修改首行表头文字、顺序或增删列。\n"
            "·「使用说明」：仅供阅读，导入时不会解析本页。\n\n"
            "二、字段与格式\n"
            "·大区经理、仓库、送货日期、节假日、品种、重量为必填；冶炼厂、仓库地址、冶炼厂地址、天气选填。\n"
            "·节假日：须手填，仅「是」（非工作日）或「否」（工作日），与业务认定一致。\n"
            "·天气：选填；不填或空则按「晴」写入；配置天气 API 时仍会按日期与地址拉取高德预报写入系统字段。\n"
            "·送货日期：可用 2026-01-15、2026/1/15、2026年1月15日、1月15日（缺省年为当年 UTC）等；"
            "亦支持 Excel 原生日期单元格或日期序列号。\n"
            "·重量：非负数字，可含小数。\n\n"
            "三、示例行\n"
            "·「导入数据」表中第 2 行起为示例（大区经理以「(示例)」开头）。导入接口会自动跳过这些行并计入 skipped；您也可在导入前自行删除。\n\n"
            "四、导入结果\n"
            "·若存在校验错误行，整批拒绝、不会部分写入；请根据接口返回的错误行号修正后重新上传。"
        )
        c = ws_help.cell(row=1, column=1, value=help_lines)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws_help.merge_cells("A1:J24")
        ws_help.row_dimensions[1].height = 360
        for col_letter in ("A", "B", "C", "D", "E", "F", "G", "H", "I", "J"):
            ws_help.column_dimensions[col_letter].width = 18

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def _normalize_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        rename_map: dict[str, str] = {}
        for c in df.columns:
            key = str(c).strip()
            if key in self.ALIAS_TO_CANONICAL:
                rename_map[c] = self.ALIAS_TO_CANONICAL[key]
        return df.rename(columns=rename_map)

    def _validate_headers(self, df: pd.DataFrame) -> None:
        cols = {str(c).strip() for c in df.columns}
        required_cn = set(self.REQUIRED_COLUMNS_CANONICAL.keys())
        optional_cn = set(self.OPTIONAL_COLUMNS_CANONICAL.keys())
        allowed = required_cn | optional_cn
        if not required_cn.issubset(cols):
            missing = required_cn - cols
            raise ValidationBusinessException(
                "表头不符合要求（须包含全部必填列含节假日，选填列仅允许仓库地址/冶炼厂地址/天气）",
                details={
                    "required": sorted(required_cn),
                    "missing": sorted(missing),
                    "unknown_extra": sorted(cols - allowed),
                },
            )
        extra = cols - allowed
        if extra:
            raise ValidationBusinessException(
                "表头不符合要求（存在未识别的列）",
                details={"required": sorted(required_cn), "optional": sorted(optional_cn), "extra": sorted(extra)},
            )

    @staticmethod
    def _parse_import_weather_cell(value: Any) -> tuple[str | None, str | None]:
        """返回 (import_weather 入库值, 错误码)。空单元格按「晴」。"""
        if value is None or (isinstance(value, float) and pd.isna(value)):
            return "晴", None
        s = str(value).strip()
        if not s:
            return "晴", None
        if len(s) > 64:
            return None, "import_weather_too_long"
        return s, None

    @staticmethod
    def _explain_import_weather_error(code: str) -> str:
        if code == "import_weather_too_long":
            return "长度不可超过 64 字符"
        return code

    def _parse_date_cell(self, value: Any) -> tuple[date | None, str | None]:
        if value is None or (isinstance(value, float) and pd.isna(value)):
            return None, "empty_date"
        if isinstance(value, datetime):
            return value.date(), None
        if isinstance(value, date):
            return value, None
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            num = float(value)
            if 1.0 <= num <= float(_EXCEL_SERIAL_MAX) and abs(num - round(num)) < 1e-9:
                try:
                    from openpyxl.utils.datetime import from_excel

                    dt = from_excel(num)
                    if isinstance(dt, datetime):
                        return dt.date(), None
                except (ValueError, OverflowError, TypeError):
                    pass
        s = str(value).strip()
        if not s:
            return None, "empty_date"
        head = s.split()[0] if " " in s else s
        m = _DATE_YMD_SEP.match(head)
        if m:
            y, mo, da = int(m.group(1)), int(m.group(2)), int(m.group(3))
            try:
                return date(y, mo, da), None
            except ValueError:
                return None, f"invalid_calendar_date:{s[:80]}"
        m_cn = _DATE_CN_YMD.match(head)
        if m_cn:
            y, mo, da = int(m_cn.group(1)), int(m_cn.group(2)), int(m_cn.group(3))
            try:
                return date(y, mo, da), None
            except ValueError:
                return None, f"invalid_calendar_date:{s[:80]}"
        m_md = _DATE_CN_MD.match(head)
        if m_md:
            mo, da = int(m_md.group(1)), int(m_md.group(2))
            y = datetime.now(timezone.utc).year
            try:
                return date(y, mo, da), None
            except ValueError:
                return None, f"invalid_calendar_date:{s[:80]}"
        parsed = pd.to_datetime(s, errors="coerce", dayfirst=False)
        if pd.isna(parsed):
            return None, f"unrecognized_date:{s[:80]}"
        ts = parsed.to_pydatetime()
        return ts.date(), None

    @staticmethod
    def _explain_date_error(code: str) -> str:
        if code == "empty_date":
            return "为空或无法识别为日期"
        if code.startswith("invalid_calendar_date:"):
            return f"日期在日历上不存在：{code.split(':', 1)[1]}"
        if code.startswith("unrecognized_date:"):
            return f"无法识别：{code.split(':', 1)[1]}"
        return code

    @staticmethod
    def _explain_weight_error(code: str) -> str:
        if code == "empty_weight":
            return "为空或非数字"
        if code.startswith("non_numeric_weight:"):
            return f"无法解析为数字：{code.split(':', 1)[1]}"
        return code

    @staticmethod
    def _parse_holiday_import_cell(value: Any) -> tuple[bool | None, str | None, str | None]:
        """返回 (cn_is_workday, cn_calendar_label, error_code)。

        「是」= 非工作日 → cn_is_workday False；「否」= 工作日 → cn_is_workday True。
        """
        if value is None or (isinstance(value, float) and pd.isna(value)):
            return None, None, "empty_holiday"
        s = str(value).strip()
        if not s:
            return None, None, "empty_holiday"
        if s == "是":
            return False, "是", None
        if s == "否":
            return True, "否", None
        return None, None, f"invalid_holiday:{s[:80]}"

    @staticmethod
    def _explain_holiday_error(code: str) -> str:
        if code == "empty_holiday":
            return "必填，仅可填「是」或「否」"
        if code.startswith("invalid_holiday:"):
            return f"仅支持「是」或「否」，当前为：{code.split(':', 1)[1]}"
        return code

    def _append_import_cell_error(
        self,
        errors: list[HistoryImportRowError],
        excel_row: int,
        column_header: str,
        message: str,
    ) -> None:
        errors.append(
            HistoryImportRowError(
                row_index=excel_row,
                excel_column=self._HEADER_TO_EXCEL_COLUMN.get(column_header),
                column_header=column_header,
                field=self.HEADER_TO_FIELD.get(column_header),
                message=message,
            )
        )

    async def _refresh_derived_delivery_fields(self, row: DeliveryRecord) -> None:
        """按日期与仓库/冶炼厂/地址尝试刷新天气 JSON（不改动导入的节假日列）。"""
        loc = " ".join(
            x
            for x in [
                (row.warehouse_address or "").strip(),
                (row.smelter_address or "").strip(),
                (row.warehouse or "").strip(),
            ]
            if x
        )
        row.weather_json = await fetch_weather_for_delivery(
            row.delivery_date,
            (row.warehouse or "").strip(),
            (row.smelter or "").strip() or None,
            loc or (row.warehouse or ""),
        )

    def _parse_weight_cell(self, value: Any) -> tuple[Decimal | None, str | None]:
        if value is None or (isinstance(value, float) and pd.isna(value)):
            return None, "empty_weight"
        if isinstance(value, (int, float)):
            return Decimal(str(value)), None
        s = str(value).strip().replace(",", "")
        if s == "":
            return None, "empty_weight"
        try:
            return Decimal(s), None
        except InvalidOperation:
            return None, f"non_numeric_weight:{s[:80]}"

    _OLE2_MAGIC = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"

    def _read_csv_dataframe(self, file_bytes: bytes, filename: str) -> pd.DataFrame:
        """读取 CSV（UTF-8 / UTF-8-BOM / GBK / GB18030），自动识别逗号或制表符分隔。"""
        text: str | None = None
        last_dec_err: Exception | None = None
        for enc in ("utf-8-sig", "utf-8", "gbk", "gb18030"):
            try:
                text = file_bytes.decode(enc)
                break
            except UnicodeDecodeError as e:
                last_dec_err = e
        if text is None:
            raise ValidationBusinessException(
                "CSV 无法按 UTF-8 或 GBK 解码，请另存为 UTF-8（带 BOM 亦可）后重试。"
            ) from last_dec_err
        try:
            return pd.read_csv(
                io.StringIO(text),
                sep=None,
                engine="python",
                dtype=object,
                skipinitialspace=True,
            )
        except Exception as e:
            logger.warning("history csv parse failed name=%s: %s", filename or "-", e)
            raise ValidationBusinessException(f"无法解析 CSV（请检查分隔符与表头）：{e}") from e

    def _read_xlsx_dataframe(self, file_bytes: bytes, filename: str) -> pd.DataFrame:
        """读取 .xlsx（OOXML zip，需 openpyxl）。"""
        if len(file_bytes) < 4 or file_bytes[:2] != b"PK":
            raise ValidationBusinessException(
                "该文件不是有效的 .xlsx（zip 包）。可改用 CSV：从 Excel「另存为 CSV UTF-8」或使用接口提供的 CSV 模板。"
            )
        try:
            return pd.read_excel(io.BytesIO(file_bytes), engine="openpyxl")
        except zipfile.BadZipFile:
            logger.warning(
                "history xlsx import: invalid zip name=%s size=%s",
                filename or "-",
                len(file_bytes),
            )
            raise ValidationBusinessException(
                "无法解析 .xlsx（文件损坏或非 Excel 工作簿）。请另存为新 .xlsx 或导出 CSV 后上传。"
            ) from None
        except Exception as e:
            logger.exception("xlsx read failed name=%s", filename or "-")
            raise ValidationBusinessException(f"无法读取 .xlsx：{e}") from e

    def _read_import_dataframe(self, file_bytes: bytes, filename: str) -> pd.DataFrame:
        """按扩展名与内容选择 CSV 或 xlsx；无扩展名时 zip 头走 xlsx，否则按 CSV 解析。"""
        fn = (filename or "").strip().lower()
        if not file_bytes:
            raise ValidationBusinessException("上传文件为空，请选择 .csv 或 .xlsx 文件")
        if len(file_bytes) >= len(self._OLE2_MAGIC) and file_bytes[: len(self._OLE2_MAGIC)] == self._OLE2_MAGIC:
            raise ValidationBusinessException(
                "检测到旧版 .xls，请另存为 .xlsx，或在 Excel 中「另存为 CSV UTF-8」后上传 .csv。"
            )
        if fn.endswith(".csv"):
            return self._read_csv_dataframe(file_bytes, filename)
        if fn.endswith(".xlsx"):
            return self._read_xlsx_dataframe(file_bytes, filename)
        if file_bytes[:2] == b"PK":
            return self._read_xlsx_dataframe(file_bytes, filename)
        return self._read_csv_dataframe(file_bytes, filename)

    async def import_excel(
        self,
        session: AsyncSession,
        file_bytes: bytes,
        filename: str,
    ) -> HistoryImportResponse:
        df = self._read_import_dataframe(file_bytes, filename)

        df = self._normalize_columns(df)
        self._validate_headers(df)

        errors: list[HistoryImportRowError] = []
        to_insert: list[DeliveryRecord] = []
        skipped = 0

        for idx, row in df.iterrows():
            excel_row = int(idx) + 2
            rm = row.get("大区经理")
            sm = row.get("冶炼厂")
            wh = row.get("仓库")
            wh_addr_cell = row.get("仓库地址")
            sm_addr_cell = row.get("冶炼厂地址")
            dv = row.get("送货日期")
            hol_cell = row.get("节假日")
            variety = row.get("品种")
            wv = row.get("重量")

            if rm is not None and str(rm).strip().startswith(self.TEMPLATE_EXAMPLE_RM_PREFIX):
                skipped += 1
                continue

            row_has_error = False
            if rm is None or str(rm).strip() == "":
                self._append_import_cell_error(errors, excel_row, "大区经理", "必填")
                row_has_error = True
            if wh is None or str(wh).strip() == "":
                self._append_import_cell_error(errors, excel_row, "仓库", "必填")
                row_has_error = True
            if variety is None or str(variety).strip() == "":
                self._append_import_cell_error(errors, excel_row, "品种", "必填")
                row_has_error = True
            sm_str: str | None = None
            if sm is not None and str(sm).strip() != "":
                sm_str = str(sm).strip()
                if len(sm_str) > 100:
                    self._append_import_cell_error(
                        errors, excel_row, "冶炼厂", "长度不可超过 100 字符"
                    )
                    row_has_error = True

            d, de = self._parse_date_cell(dv)
            if de:
                self._append_import_cell_error(
                    errors, excel_row, "送货日期", self._explain_date_error(de)
                )
                row_has_error = True

            w, we = self._parse_weight_cell(wv)
            if we:
                self._append_import_cell_error(
                    errors, excel_row, "重量", self._explain_weight_error(we)
                )
                row_has_error = True
            if w is not None and w < 0:
                self._append_import_cell_error(errors, excel_row, "重量", "不可为负")
                row_has_error = True

            cn_wd, cn_lab, he = self._parse_holiday_import_cell(hol_cell)
            if he:
                self._append_import_cell_error(
                    errors, excel_row, "节假日", self._explain_holiday_error(he)
                )
                row_has_error = True

            wa: str | None = None
            if wh_addr_cell is not None and str(wh_addr_cell).strip():
                wa = str(wh_addr_cell).strip()
                if len(wa) > 512:
                    self._append_import_cell_error(
                        errors, excel_row, "仓库地址", "长度不可超过 512 字符"
                    )
                    row_has_error = True
            sa: str | None = None
            if sm_addr_cell is not None and str(sm_addr_cell).strip():
                sa = str(sm_addr_cell).strip()
                if len(sa) > 512:
                    self._append_import_cell_error(
                        errors, excel_row, "冶炼厂地址", "长度不可超过 512 字符"
                    )
                    row_has_error = True

            wt_cell = row.get("天气")
            iwt, wte = self._parse_import_weather_cell(wt_cell)
            if wte:
                self._append_import_cell_error(
                    errors, excel_row, "天气", self._explain_import_weather_error(wte)
                )
                row_has_error = True

            if row_has_error:
                continue

            assert d is not None and w is not None and cn_wd is not None and cn_lab is not None and iwt is not None
            loc = " ".join(x for x in [wa or "", sa or "", str(wh).strip()] if x)
            weather_json = await fetch_weather_for_delivery(
                d,
                str(wh).strip(),
                sm_str,
                loc or str(wh).strip(),
            )
            to_insert.append(
                DeliveryRecord(
                    regional_manager=str(rm).strip(),
                    smelter=sm_str,
                    warehouse=str(wh).strip(),
                    warehouse_address=wa,
                    smelter_address=sa,
                    delivery_date=d,
                    product_variety=str(variety).strip(),
                    weight=w,
                    cn_is_workday=cn_wd,
                    cn_calendar_label=cn_lab,
                    weather_json=weather_json,
                    import_weather=iwt,
                )
            )

        if errors:
            raise ValidationBusinessException(
                "导入失败：存在错误行，已整批拒绝",
                details={"errors": [e.model_dump() for e in errors]},
            )

        for rec in to_insert:
            session.add(rec)

        inserted = len(to_insert)
        logger.info("history import finished inserted=%s skipped=%s", inserted, skipped)
        return HistoryImportResponse(inserted=inserted, skipped=skipped, errors=[])

    async def list_records(
        self,
        session: AsyncSession,
        q: HistoryQueryParams,
    ) -> HistoryListResponse:
        stmt = select(DeliveryRecord)
        count_stmt = select(func.count()).select_from(DeliveryRecord)

        rms = list(q.regional_managers)
        if not rms and q.regional_manager:
            rms = [q.regional_manager]
        if rms:
            stmt = stmt.where(DeliveryRecord.regional_manager.in_(rms))
            count_stmt = count_stmt.where(DeliveryRecord.regional_manager.in_(rms))

        whs = list(q.warehouses)
        if not whs and q.warehouse:
            whs = [q.warehouse]
        if whs:
            stmt = stmt.where(DeliveryRecord.warehouse.in_(whs))
            count_stmt = count_stmt.where(DeliveryRecord.warehouse.in_(whs))

        vars_ = list(q.product_varieties)
        if not vars_ and q.product_variety:
            vars_ = [q.product_variety]
        if vars_:
            stmt = stmt.where(DeliveryRecord.product_variety.in_(vars_))
            count_stmt = count_stmt.where(DeliveryRecord.product_variety.in_(vars_))

        sms = list(q.smelters)
        if not sms and q.smelter:
            sms = [q.smelter]
        if sms:
            stmt = stmt.where(DeliveryRecord.smelter.in_(sms))
            count_stmt = count_stmt.where(DeliveryRecord.smelter.in_(sms))

        if q.date_from:
            stmt = stmt.where(DeliveryRecord.delivery_date >= q.date_from)
            count_stmt = count_stmt.where(DeliveryRecord.delivery_date >= q.date_from)
        if q.date_to:
            stmt = stmt.where(DeliveryRecord.delivery_date <= q.date_to)
            count_stmt = count_stmt.where(DeliveryRecord.delivery_date <= q.date_to)

        total_res = await session.execute(count_stmt)
        total = int(total_res.scalar_one())

        stmt = stmt.order_by(DeliveryRecord.delivery_date.desc(), DeliveryRecord.id.desc())
        offset = (q.page - 1) * q.page_size
        stmt = stmt.offset(offset).limit(q.page_size)
        res = await session.execute(stmt)
        rows = res.scalars().all()
        items = [DeliveryRecordRead.model_validate(r, from_attributes=True) for r in rows]
        return HistoryListResponse(total=total, page=q.page, page_size=q.page_size, items=items)

    def _history_filter_clauses(self, q: HistoryQueryParams) -> list[Any]:
        """与 list_records 相同的筛选条件（不含分页）。"""
        clauses: list[Any] = []
        rms = list(q.regional_managers)
        if not rms and q.regional_manager:
            rms = [q.regional_manager]
        if rms:
            clauses.append(DeliveryRecord.regional_manager.in_(rms))

        whs = list(q.warehouses)
        if not whs and q.warehouse:
            whs = [q.warehouse]
        if whs:
            clauses.append(DeliveryRecord.warehouse.in_(whs))

        vars_ = list(q.product_varieties)
        if not vars_ and q.product_variety:
            vars_ = [q.product_variety]
        if vars_:
            clauses.append(DeliveryRecord.product_variety.in_(vars_))

        sms = list(q.smelters)
        if not sms and q.smelter:
            sms = [q.smelter]
        if sms:
            clauses.append(DeliveryRecord.smelter.in_(sms))

        if q.date_from:
            clauses.append(DeliveryRecord.delivery_date >= q.date_from)
        if q.date_to:
            clauses.append(DeliveryRecord.delivery_date <= q.date_to)
        return clauses

    async def statistics(
        self,
        session: AsyncSession,
        q: HistoryQueryParams,
        *,
        top_n: int = 200,
    ) -> HistoryStatsResponse:
        clauses = self._history_filter_clauses(q)
        wc = and_(*clauses) if clauses else True

        cnt_stmt = select(func.count(DeliveryRecord.id)).where(wc)
        sum_stmt = select(func.coalesce(func.sum(DeliveryRecord.weight), 0)).where(wc)
        total = int((await session.execute(cnt_stmt)).scalar_one() or 0)
        tw = (await session.execute(sum_stmt)).scalar_one()
        total_weight = Decimal(str(tw)) if tw is not None else Decimal("0")

        async def _bucket_rows(key_col: Any) -> list[HistoryStatsBucket]:
            stmt = (
                select(
                    key_col,
                    func.count(DeliveryRecord.id),
                    func.coalesce(func.sum(DeliveryRecord.weight), 0),
                )
                .where(wc)
                .group_by(key_col)
                .order_by(desc(func.coalesce(func.sum(DeliveryRecord.weight), 0)))
                .limit(top_n)
            )
            res = await session.execute(stmt)
            out: list[HistoryStatsBucket] = []
            for label, c, w in res.all():
                out.append(
                    HistoryStatsBucket(
                        key=str(label),
                        record_count=int(c or 0),
                        total_weight=Decimal(str(w or 0)),
                    )
                )
            return out

        by_wh = await _bucket_rows(DeliveryRecord.warehouse)
        by_var = await _bucket_rows(DeliveryRecord.product_variety)
        by_rm = await _bucket_rows(DeliveryRecord.regional_manager)
        return HistoryStatsResponse(
            total_records=total,
            total_weight=total_weight,
            date_from=q.date_from,
            date_to=q.date_to,
            by_warehouse=by_wh,
            by_product_variety=by_var,
            by_regional_manager=by_rm,
        )

    async def update_record(
        self,
        session: AsyncSession,
        record_id: int,
        payload: DeliveryRecordUpdate,
    ) -> DeliveryRecordRead | None:
        patch = payload.model_dump(exclude_unset=True)
        if not patch:
            raise ValidationBusinessException("未提供任何可更新字段")
        stmt = select(DeliveryRecord).where(DeliveryRecord.id == record_id)
        row = (await session.execute(stmt)).scalar_one_or_none()
        if row is None:
            return None
        if "regional_manager" in patch and not str(patch["regional_manager"]).strip():
            raise ValidationBusinessException("大区经理不可为空")
        if "warehouse" in patch and not str(patch["warehouse"]).strip():
            raise ValidationBusinessException("仓库不可为空")
        if "product_variety" in patch and not str(patch["product_variety"]).strip():
            raise ValidationBusinessException("品种不可为空")
        for k, v in patch.items():
            setattr(row, k, v)
        if "cn_calendar_label" in patch and row.cn_calendar_label in ("是", "否"):
            row.cn_is_workday = row.cn_calendar_label == "否"
        weather_keys = {"delivery_date", "warehouse_address", "smelter_address", "warehouse", "smelter"}
        if patch.keys() & weather_keys:
            await self._refresh_derived_delivery_fields(row)
        await session.flush()
        return DeliveryRecordRead.model_validate(row, from_attributes=True)

    async def batch_delete(self, session: AsyncSession, ids: list[int]) -> int:
        if not ids:
            return 0
        stmt = delete(DeliveryRecord).where(DeliveryRecord.id.in_(ids))
        res = await session.execute(stmt)
        return int(res.rowcount or 0)

    async def purge_all_delivery_records(self, session: AsyncSession) -> int:
        """删除送货历史表全部行（慎用）。"""
        res = await session.execute(delete(DeliveryRecord))
        return int(res.rowcount or 0)


def get_history_service() -> HistoryService:
    return HistoryService()
