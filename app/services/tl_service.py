"""
TL比价模块服务层
负责仓库、冶炼厂、品类、比价、运费、价格表、品类映射等数据库操作
"""
import hashlib
import io
import json
import logging
import os
import uuid
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from app.config import UPLOAD_DIR
from app.database import get_conn
from app.models.tl import OPTIMAL_PRICE_BASIS_ALLOWED
from app.quote_price_sources import (
    API_KEY_TO_DB,
    merge_sources_after_fill,
    normalize_client_sources,
    SOURCE_DERIVED,
    SOURCE_ORIGINAL,
)
from app.price_tax_utils import (
    derive_net_and_vat_from_quote_row,
    derive_vat_prices_from_stated_price,
    fill_vat_from_exclusive_net,
    inclusive_from_net,
    merge_factory_rates,
    net_from_inclusive,
    parse_price_basis_from_remark,
)
from app.services.vlm_extractor_service import QwenVLFullExtractor, VLMConfig

logger = logging.getLogger(__name__)


def _comparison_quote_calendar_date() -> date:
    """
    比价使用的「当天」报价日期（仅取 quote_details 中该日的行，不取历史上传）。
    默认按 Asia/Shanghai 日历日；可通过环境变量 QUOTE_COMPARISON_TZ 设为其它 IANA 时区（如 UTC）。
    """
    tz_name = (os.getenv("QUOTE_COMPARISON_TZ") or "Asia/Shanghai").strip() or "Asia/Shanghai"
    try:
        from zoneinfo import ZoneInfo

        return datetime.now(ZoneInfo(tz_name)).date()
    except Exception:
        logger.warning(
            "QUOTE_COMPARISON_TZ=%r 无效，比价日期回退为服务器本地当天", tz_name
        )
        return date.today()


def _unit_for_optimal_price_basis(
    basis: str,
    breakdown: Optional[Tuple[float, float, float, float]],
    qrow: Optional[Dict[str, Optional[float]]],
) -> Optional[float]:
    """
    最优价用的单价（元/吨）：base/1pct/3pct/13pct 来自统一反推 breakdown；
    普票、反向发票取表中对应列。
    """
    if basis == "base":
        return breakdown[0] if breakdown else None
    if basis == "1pct":
        return breakdown[1] if breakdown else None
    if basis == "3pct":
        return breakdown[2] if breakdown else None
    if basis == "13pct":
        return breakdown[3] if breakdown else None
    if basis == "normal_invoice":
        if not qrow:
            return None
        v = qrow.get("price_normal_invoice")
        return float(v) if v is not None else None
    if basis == "reverse_invoice":
        if not qrow:
            return None
        v = qrow.get("price_reverse_invoice")
        return float(v) if v is not None else None
    return None


class PurchaseSuggestionLLMError(Exception):
    """采购建议接口调用上游大模型失败（由路由映射为 HTTP 502）。"""


PRICE_TABLE_UPLOAD_DIR = Path(UPLOAD_DIR) / "price_tables"
PRICE_TABLE_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)


def _cell_json(v: Any) -> Any:
    if v is None:
        return None
    if isinstance(v, Decimal):
        return float(v)
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(v, date):
        return v.isoformat()
    return v


def _json_cell_to_dict(val: Any) -> Optional[Dict[str, Any]]:
    """解析库表 JSON 列（或已解析的 dict）为字典。"""
    if val is None:
        return None
    if isinstance(val, dict):
        return val
    if isinstance(val, (bytes, bytearray)):
        val = val.decode("utf-8")
    if isinstance(val, str):
        s = val.strip()
        if not s:
            return None
        return json.loads(s)
    return None


def _apply_factory_tax_rates_to_quote_item(
    item: Dict[str, Any],
    tax_by_fid: Dict[int, Dict[str, float]],
) -> bool:
    """
    确认写入前：按系统为冶炼厂保存的税率（factory_tax_rates ∪ 默认）统一落库。

    - 若识别/前端给出的是**不含税基准**（「价格」有值）→ 用合并税率**正算**含1%/3%/13%价。
    - 若仅有**含税价**列（1%/3%/13% 之一）→ 用对应税率**反算**不含税基准，再**正算**三档含税（顺序：优先 13% 列 → 3% → 1%）。

    与图片识别一致：图上可能是基准也可能是含税；最终以本函数 + 系统税率为准写入 quote_details。
    """
    fid = item.get("冶炼厂id")
    if fid is None:
        return False
    merged = merge_factory_rates(tax_by_fid.get(int(fid)))

    net: Optional[float] = None
    if item.get("价格") is not None:
        net = float(item["价格"])
    elif item.get("价格_13pct增值税") is not None and "13pct" in merged:
        net = net_from_inclusive(float(item["价格_13pct增值税"]), merged["13pct"])
    elif item.get("价格_3pct增值税") is not None and "3pct" in merged:
        net = net_from_inclusive(float(item["价格_3pct增值税"]), merged["3pct"])
    elif item.get("价格_1pct增值税") is not None and "1pct" in merged:
        net = net_from_inclusive(float(item["价格_1pct增值税"]), merged["1pct"])

    if net is None:
        return False

    item["价格"] = round(float(net), 2)
    f1, f3, f13 = fill_vat_from_exclusive_net(float(item["价格"]), merged)
    item["价格_1pct增值税"] = f1
    item["价格_3pct增值税"] = f3
    item["价格_13pct增值税"] = f13
    return True


class TLService:

    # ==================== 接口0：添加仓库 ====================

    def add_warehouse(self, name: str) -> Dict[str, Any]:
        try:
            with get_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        "SELECT id FROM dict_warehouses WHERE name = %s",
                        (name,),
                    )
                    row = cur.fetchone()
                    if row:
                        return {"code": 200, "msg": "仓库已存在", "仓库id": row[0], "新建": False}
                    cur.execute(
                        "INSERT INTO dict_warehouses (name, is_active) VALUES (%s, 1)",
                        (name,),
                    )
                    return {"code": 200, "msg": "仓库新建成功", "仓库id": cur.lastrowid, "新建": True}
        except Exception as e:
            logger.error(f"添加仓库失败: {e}")
            raise

    # ==================== 接口0b：新建冶炼厂 ====================

    def add_smelter(self, name: str) -> Dict[str, Any]:
        try:
            with get_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        "SELECT id, is_active FROM dict_factories WHERE name = %s",
                        (name,),
                    )
                    row = cur.fetchone()
                    if row:
                        smelter_id, is_active = row
                        if is_active == 1:
                            return {"code": 200, "msg": "冶炼厂已存在", "冶炼厂id": smelter_id, "新建": False}
                        cur.execute(
                            "UPDATE dict_factories SET is_active = 1 WHERE id = %s",
                            (smelter_id,),
                        )
                        return {"code": 200, "msg": "冶炼厂已恢复启用", "冶炼厂id": smelter_id, "新建": False}

                    cur.execute(
                        "INSERT INTO dict_factories (name, is_active) VALUES (%s, 1)",
                        (name,),
                    )
                    return {"code": 200, "msg": "冶炼厂新建成功", "冶炼厂id": cur.lastrowid, "新建": True}
        except Exception as e:
            logger.error(f"新建冶炼厂失败: {e}")
            raise

    # ==================== 接口1：获取仓库列表 ====================

    def get_warehouses(self, keyword: Optional[str] = None) -> List[Dict[str, Any]]:
        try:
            conditions = ["is_active = 1"]
            params: List[Any] = []
            if keyword is not None and str(keyword).strip():
                conditions.append("name LIKE %s")
                params.append(f"%{str(keyword).strip()}%")
            where_sql = " AND ".join(conditions)
            with get_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        f"SELECT id AS `仓库id`, name AS `仓库名` "
                        f"FROM dict_warehouses WHERE {where_sql} "
                        "ORDER BY id",
                        tuple(params),
                    )
                    columns = [desc[0] for desc in cur.description]
                    rows = cur.fetchall()
                    return [dict(zip(columns, row)) for row in rows]
        except Exception as e:
            logger.error(f"获取仓库列表失败: {e}")
            raise

    # ==================== 接口1b：修改仓库 ====================

    def update_warehouse(
        self,
        warehouse_id: int,
        name: Optional[str] = None,
        is_active: Optional[bool] = None,
    ) -> Dict[str, Any]:
        if name is None and is_active is None:
            raise ValueError("至少需要提供一个待修改字段：仓库名 或 is_active")

        try:
            with get_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute("SELECT id FROM dict_warehouses WHERE id = %s", (warehouse_id,))
                    if not cur.fetchone():
                        raise ValueError(f"仓库 id={warehouse_id} 不存在")

                    updates = []
                    params: List[Any] = []

                    if name is not None:
                        cur.execute(
                            "SELECT id FROM dict_warehouses WHERE name = %s AND id <> %s",
                            (name, warehouse_id),
                        )
                        if cur.fetchone():
                            raise ValueError(f"仓库名 '{name}' 已存在")
                        updates.append("name = %s")
                        params.append(name)

                    if is_active is not None:
                        updates.append("is_active = %s")
                        params.append(1 if is_active else 0)

                    params.append(warehouse_id)
                    cur.execute(
                        f"UPDATE dict_warehouses SET {', '.join(updates)} WHERE id = %s",
                        tuple(params),
                    )

            return {"code": 200, "msg": "仓库信息修改成功"}
        except ValueError:
            raise
        except Exception as e:
            logger.error(f"修改仓库失败: {e}")
            raise

    # ==================== 接口1c：删除仓库（软删除） ====================

    def delete_warehouse(self, warehouse_id: int) -> Dict[str, Any]:
        try:
            with get_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        "SELECT id FROM dict_warehouses WHERE id = %s AND is_active = 1",
                        (warehouse_id,),
                    )
                    if not cur.fetchone():
                        raise ValueError(f"仓库 id={warehouse_id} 不存在或已删除")

                    cur.execute(
                        "UPDATE dict_warehouses SET is_active = 0 WHERE id = %s",
                        (warehouse_id,),
                    )
            return {"code": 200, "msg": "仓库已删除"}
        except ValueError:
            raise
        except Exception as e:
            logger.error(f"删除仓库失败: {e}")
            raise

    # ==================== 接口2：获取冶炼厂列表 ====================

    def get_smelters(self) -> List[Dict[str, Any]]:
        try:
            with get_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        "SELECT id AS `冶炼厂id`, name AS `冶炼厂` "
                        "FROM dict_factories "
                        "WHERE is_active = 1 "
                        "ORDER BY id"
                    )
                    columns = [desc[0] for desc in cur.description]
                    rows = cur.fetchall()
                    return [dict(zip(columns, row)) for row in rows]
        except Exception as e:
            logger.error(f"获取冶炼厂列表失败: {e}")
            raise

    # ==================== 接口2b：修改冶炼厂 ====================

    def update_smelter(
        self,
        smelter_id: int,
        name: Optional[str] = None,
        is_active: Optional[bool] = None,
    ) -> Dict[str, Any]:
        if name is None and is_active is None:
            raise ValueError("至少需要提供一个待修改字段：冶炼厂名 或 is_active")

        try:
            with get_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute("SELECT id FROM dict_factories WHERE id = %s", (smelter_id,))
                    if not cur.fetchone():
                        raise ValueError(f"冶炼厂 id={smelter_id} 不存在")

                    updates = []
                    params: List[Any] = []

                    if name is not None:
                        cur.execute(
                            "SELECT id FROM dict_factories WHERE name = %s AND id <> %s",
                            (name, smelter_id),
                        )
                        if cur.fetchone():
                            raise ValueError(f"冶炼厂名 '{name}' 已存在")
                        updates.append("name = %s")
                        params.append(name)

                    if is_active is not None:
                        updates.append("is_active = %s")
                        params.append(1 if is_active else 0)

                    params.append(smelter_id)
                    cur.execute(
                        f"UPDATE dict_factories SET {', '.join(updates)} WHERE id = %s",
                        tuple(params),
                    )

            return {"code": 200, "msg": "冶炼厂信息修改成功"}
        except ValueError:
            raise
        except Exception as e:
            logger.error(f"修改冶炼厂失败: {e}")
            raise

    # ==================== 接口2c：删除冶炼厂（软删除） ====================

    def delete_smelter(self, smelter_id: int) -> Dict[str, Any]:
        try:
            with get_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        "SELECT id FROM dict_factories WHERE id = %s AND is_active = 1",
                        (smelter_id,),
                    )
                    if not cur.fetchone():
                        raise ValueError(f"冶炼厂 id={smelter_id} 不存在或已删除")

                    cur.execute(
                        "UPDATE dict_factories SET is_active = 0 WHERE id = %s",
                        (smelter_id,),
                    )
            return {"code": 200, "msg": "冶炼厂已删除"}
        except ValueError:
            raise
        except Exception as e:
            logger.error(f"删除冶炼厂失败: {e}")
            raise

    # ==================== 接口3：获取品类列表 ====================

    def get_categories(self) -> List[Dict[str, Any]]:
        try:
            with get_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        "SELECT category_id AS `品类id`, "
                        "GROUP_CONCAT(name ORDER BY row_id SEPARATOR '、') AS `品类名` "
                        "FROM dict_categories "
                        "WHERE is_active = 1 "
                        "GROUP BY category_id "
                        "ORDER BY category_id"
                    )
                    columns = [desc[0] for desc in cur.description]
                    rows = cur.fetchall()
                    return [dict(zip(columns, row)) for row in rows]
        except Exception as e:
            logger.error(f"获取品类列表失败: {e}")
            raise

    # ==================== 接口3b：上传品种 ====================

    def upload_variety(self, items: List[Dict[str, Any]]) -> Dict[str, Any]:
        """
        批量维护品种（dict_categories）：不存在则新建分组并 is_main=1；
        名称已存在且启用则跳过；已存在但停用则恢复启用。
        """
        if not items:
            raise ValueError("品种数据不能为空")

        seen: set[str] = set()
        names: List[str] = []
        for item in items:
            raw = item.get("品种名")
            if raw is None:
                continue
            n = str(raw).strip()
            if not n:
                continue
            if len(n) > 50:
                raise ValueError(f"品种名长度不能超过50字符: {n[:30]}…")
            if n in seen:
                continue
            seen.add(n)
            names.append(n)

        if not names:
            raise ValueError("无有效的品种名")

        created = existed = reactivated = 0
        try:
            with get_conn() as conn:
                with conn.cursor() as cur:
                    for n in names:
                        cur.execute(
                            "SELECT row_id, category_id, is_active "
                            "FROM dict_categories WHERE name = %s",
                            (n,),
                        )
                        row = cur.fetchone()
                        if row:
                            _rid, _cid, is_active = row
                            if is_active == 1:
                                existed += 1
                            else:
                                cur.execute(
                                    "UPDATE dict_categories SET is_active = 1 WHERE row_id = %s",
                                    (_rid,),
                                )
                                reactivated += 1
                        else:
                            cur.execute(
                                "SELECT COALESCE(MAX(category_id), 0) + 1 FROM dict_categories"
                            )
                            new_cat_id = cur.fetchone()[0]
                            cur.execute(
                                "INSERT INTO dict_categories "
                                "(category_id, name, is_main, is_active) "
                                "VALUES (%s, %s, 1, 1)",
                                (new_cat_id, n),
                            )
                            created += 1

            parts = []
            if created:
                parts.append(f"新建 {created} 个")
            if existed:
                parts.append(f"已存在 {existed} 个")
            if reactivated:
                parts.append(f"恢复启用 {reactivated} 个")
            msg = "、".join(parts) if parts else "无变更"
            return {
                "code": 200,
                "msg": f"品种已处理：{msg}",
                "新建": created,
                "已存在": existed,
                "恢复启用": reactivated,
            }
        except ValueError:
            raise
        except Exception as e:
            logger.error(f"上传品种失败: {e}")
            raise

    # ==================== 接口4：获取比价表 ====================
    def get_comparison(
        self,
        warehouse_ids: List[int],
        smelter_ids: List[int],
        category_ids: List[int],
        price_type: Optional[str] = None,
        tons: float = 1.0,
        optimal_basis_list: Optional[List[str]] = None,
        optimal_sort_basis: Optional[str] = None,
        quote_date_str: Optional[str] = None,
    ) -> Dict[str, Any]:
        """
        price_type: 目标税率类型，None=普通价, 1pct/3pct/13pct/normal_invoice/reverse_invoice
        吨数 t: 报价按元/吨；**总运费** = 运费单价（元/吨）× t，即 **运费×吨数**。
        **展示用「报价」**：按所选 price_type 折合为不含税（元/吨）后写入 `报价`；
        **`报价金额`** = **报价×吨数**（无报价为 `null`）；**`利润`** = **报价金额 − 总运费**（即 **报价×吨数 − 运费×吨数**）。
        前端最终比价、明细排序与 **`冶炼厂利润排行`** 均以该 **`利润`**（及所选最优价口径）为准。
        **最优价各口径利润**=该口径下元/吨单价×t−总运费（与主利润同一套总运费）。
        同时按表中已有列统一反推 `基准价`（不含税）、`含1%税价`、`含3%税价`（与 OCR 按税点入库、再换算一致）；
        `利润_基准`=基准价×t−总运费，`利润_含3%`=含3%税价×t−总运费。
        **最优价各口径利润**：由 optimal_basis_list 指定（如 base、1pct、3pct、13pct、普票列等），每条明细返回 `最优价各口径利润` 字典；
        明细与冶炼厂排行按 optimal_sort_basis（默认列表首项）对应利润从高到低排序。
        取价逻辑（按优先级）：
          1. 报价表中直接有对应 price_type 的价格 → 直接使用
          2. 有不含税 unit_price（基准价）+ 税率表 → 目标含税价 = unit_price × (1+税率)
          3. 目标是普通价(unit_price) 但列空 → 由已知含1%/3%/13%价反算不含税基准
          4. 仅有某一档含税价 → 先反算不含税，再换算到目标税率
          5. 以上均无 → None，返回 price_source="unavailable"

        **报价日期**：
        - 若传入 `quote_date_str`（YYYY-MM-DD）：只使用该日的 `quote_details`。
        - 否则：对每个 (冶炼厂, 品种名) 取 **quote_date ≤ 比价日历日**（默认 `Asia/Shanghai`，见 `QUOTE_COMPARISON_TZ`）
          的 **最近一条** 报价，避免「确认报价不是今天」时比价整表无单价。
        """
        if not warehouse_ids or not smelter_ids or not category_ids:
            return {
                "明细": [],
                "冶炼厂利润排行": [],
                "最优价排序口径": (optimal_sort_basis or (optimal_basis_list or ["3pct"])[0]),
            }

        bases = list(optimal_basis_list or ["3pct"])
        sort_basis = optimal_sort_basis if optimal_sort_basis is not None else bases[0]
        for b in bases:
            if b not in OPTIMAL_PRICE_BASIS_ALLOWED:
                raise ValueError(
                    f"不支持的最优价计税口径: {b!r}，允许：{sorted(OPTIMAL_PRICE_BASIS_ALLOWED)}"
                )
        if sort_basis not in bases:
            raise ValueError(
                f"最优价排序口径 {sort_basis!r} 须在最优价计税口径列表中，当前为 {bases}"
            )

        # price_type → (quote_details列名, 展示名)
        PRICE_COL_MAP = {
            None:             ("unit_price",            "普通价"),
            "1pct":           ("price_1pct_vat",        "1%增值税"),
            "3pct":           ("price_3pct_vat",        "3%增值税"),
            "13pct":          ("price_13pct_vat",       "13%增值税"),
            "normal_invoice": ("price_normal_invoice",  "普通发票"),
            "reverse_invoice":("price_reverse_invoice", "反向发票"),
        }
        # 仅以下三种有税率换算意义
        VAT_TAX_TYPE_MAP = {"1pct": "1pct", "3pct": "3pct", "13pct": "13pct"}

        if price_type not in PRICE_COL_MAP:
            raise ValueError(f"不支持的 price_type: {price_type}")

        target_col, price_type_name = PRICE_COL_MAP[price_type]
        target_tax = VAT_TAX_TYPE_MAP.get(price_type)  # None 表示不需要税率换算

        try:
            with get_conn() as conn:
                with conn.cursor() as cur:
                    wh_ph = ",".join(["%s"] * len(warehouse_ids))
                    sm_ph = ",".join(["%s"] * len(smelter_ids))
                    cat_ph = ",".join(["%s"] * len(category_ids))

                    # 品类主名称（用于展示）
                    cur.execute(
                        f"SELECT DISTINCT category_id, "
                        f"COALESCE(MAX(CASE WHEN is_main=1 THEN name END), MAX(name)) AS cat_name "
                        f"FROM dict_categories "
                        f"WHERE category_id IN ({cat_ph}) AND is_active = 1 "
                        f"GROUP BY category_id",
                        tuple(category_ids),
                    )
                    cat_map: Dict[int, str] = {row[0]: row[1] for row in cur.fetchall()}

                    # 最新运费
                    cur.execute(
                        f"""
                        SELECT dw.id, dw.name, df.id, df.name, fr.price_per_ton
                        FROM freight_rates fr
                        JOIN dict_warehouses dw ON fr.warehouse_id = dw.id
                        JOIN dict_factories  df ON fr.factory_id  = df.id
                        WHERE dw.id IN ({wh_ph})
                          AND df.id IN ({sm_ph})
                          AND fr.effective_date = (
                              SELECT MAX(fr2.effective_date)
                              FROM freight_rates fr2
                              WHERE fr2.factory_id  = fr.factory_id
                                AND fr2.warehouse_id = fr.warehouse_id
                          )
                        """,
                        tuple(warehouse_ids) + tuple(smelter_ids),
                    )
                    freight_map: Dict[tuple, tuple] = {}
                    for wid, wname, fid, fname, freight in cur.fetchall():
                        freight_map[(wid, fid)] = (wname, fname, freight)

                    # category_id → 品类名称列表（用于匹配价格表）
                    cur.execute(
                        f"SELECT category_id, name FROM dict_categories "
                        f"WHERE category_id IN ({cat_ph}) AND is_active = 1",
                        tuple(category_ids),
                    )
                    cat_id_to_names: Dict[int, List[str]] = {}
                    for cat_id, name in cur.fetchall():
                        n = str(name).strip()
                        if not n:
                            continue
                        lst = cat_id_to_names.setdefault(cat_id, [])
                        if n not in lst:
                            lst.append(n)

                    if not cat_id_to_names:
                        return {
                            "明细": [],
                            "冶炼厂利润排行": [],
                            "最优价排序口径": sort_basis,
                        }

                    # 所有品类名称（去重，与 quote_details 用 TRIM 后匹配）
                    all_cat_names: List[str] = []
                    _seen_cn: set = set()
                    for names in cat_id_to_names.values():
                        for n in names:
                            if n not in _seen_cn:
                                _seen_cn.add(n)
                                all_cat_names.append(n)
                    cn_ph = ",".join(["%s"] * len(all_cat_names))

                    # 税率表：{factory_id: {tax_type: rate}}
                    cur.execute(
                        f"SELECT factory_id, tax_type, tax_rate "
                        f"FROM factory_tax_rates "
                        f"WHERE factory_id IN ({sm_ph})",
                        tuple(smelter_ids),
                    )
                    tax_rate_map: Dict[int, Dict[str, float]] = {}
                    for fid, ttype, rate in cur.fetchall():
                        tax_rate_map.setdefault(fid, {})[ttype] = float(rate)

                    quote_day = _comparison_quote_calendar_date()
                    if quote_date_str is not None and str(quote_date_str).strip():
                        try:
                            exact_qd = date.fromisoformat(str(quote_date_str).strip())
                        except (ValueError, TypeError):
                            raise ValueError(
                                f"报价日期 格式不正确: {quote_date_str}，应为 YYYY-MM-DD"
                            )
                        cur.execute(
                            f"""
                            SELECT factory_id, TRIM(category_name) AS category_name,
                                   unit_price, price_1pct_vat, price_3pct_vat, price_13pct_vat,
                                   price_normal_invoice, price_reverse_invoice
                            FROM quote_details
                            WHERE factory_id IN ({sm_ph})
                              AND TRIM(category_name) IN ({cn_ph})
                              AND quote_date = %s
                            """,
                            tuple(smelter_ids) + tuple(all_cat_names) + (exact_qd,),
                        )
                    else:
                        cur.execute(
                            f"""
                            SELECT qd.factory_id, TRIM(qd.category_name) AS category_name,
                                   qd.unit_price, qd.price_1pct_vat, qd.price_3pct_vat,
                                   qd.price_13pct_vat,
                                   qd.price_normal_invoice, qd.price_reverse_invoice
                            FROM quote_details qd
                            INNER JOIN (
                                SELECT factory_id, TRIM(category_name) AS cname, MAX(quote_date) AS mx
                                FROM quote_details
                                WHERE factory_id IN ({sm_ph})
                                  AND TRIM(category_name) IN ({cn_ph})
                                  AND quote_date <= %s
                                GROUP BY factory_id, TRIM(category_name)
                            ) latest ON latest.factory_id = qd.factory_id
                                AND TRIM(qd.category_name) = latest.cname
                                AND latest.mx = qd.quote_date
                            WHERE qd.factory_id IN ({sm_ph})
                              AND TRIM(qd.category_name) IN ({cn_ph})
                            """,
                            tuple(smelter_ids)
                            + tuple(all_cat_names)
                            + (quote_day,)
                            + tuple(smelter_ids)
                            + tuple(all_cat_names),
                        )
                    # raw_price_map: {(factory_id, category_name): {col: value}}
                    col_names = ["unit_price", "price_1pct_vat", "price_3pct_vat",
                                 "price_13pct_vat", "price_normal_invoice", "price_reverse_invoice"]
                    raw_price_map: Dict[tuple, Dict[str, Optional[float]]] = {}
                    name_to_cat_id: Dict[str, int] = {}
                    for row in cur.fetchall():
                        fid_r, cat_name = row[0], str(row[1]).strip() if row[1] is not None else ""
                        if not cat_name:
                            continue
                        raw_price_map[(fid_r, cat_name)] = {
                            col: (float(v) if v is not None else None)
                            for col, v in zip(col_names, row[2:])
                        }
                        for cat_id, names in cat_id_to_names.items():
                            if cat_name in names:
                                name_to_cat_id[cat_name] = cat_id
                                break

            # 换算逻辑（纯 Python，连接已关闭）
            # col → tax_type 的对应关系，用于反算不含税价
            COL_TO_TAX: Dict[str, str] = {
                "price_1pct_vat": "1pct",
                "price_3pct_vat": "3pct",
                "price_13pct_vat": "13pct",
            }

            def resolve_price(fid: int, cat_id: int) -> Tuple[Optional[float], str]:
                """
                返回 (price, source)
                source: "direct" | "calc_from_base" | "calc_from_other_vat" | "unavailable"
                """
                # 找该 category_id 下的所有品类名称，取第一个有报价的
                cat_names = cat_id_to_names.get(cat_id, [])
                for cat_name in cat_names:
                    prices = raw_price_map.get((fid, cat_name), {})
                    if not prices or not any(
                        v is not None for v in prices.values()
                    ):
                        continue

                    rates = tax_rate_map.get(fid, {})
                    merged = merge_factory_rates(rates)

                    # 1. 直接有目标列
                    direct = prices.get(target_col)
                    if direct is not None:
                        return direct, "direct"

                    # 2. 不含税 unit_price → 目标税率含税价
                    if target_tax and prices.get("unit_price") is not None and target_tax in merged:
                        up = float(prices["unit_price"])
                        calc = inclusive_from_net(up, merged[target_tax])
                        return calc, "calc_from_base"

                    # 3. 目标为不含税基准，由已知含税价反算
                    if target_col == "unit_price":
                        for col, src_tax in COL_TO_TAX.items():
                            known_price = prices.get(col)
                            if known_price is not None and src_tax in merged:
                                net = net_from_inclusive(float(known_price), merged[src_tax])
                                return round(net, 2), f"calc_from_{src_tax}"
                        # 与 derive_net_and_vat_from_quote_row 一致：仅有普票/反向发票列时按不含税理解
                        for col in ("price_normal_invoice", "price_reverse_invoice"):
                            v = prices.get(col)
                            if v is not None:
                                return float(v), f"direct_{col}"

                    # 4. 从某一档含税价反算不含税，再换算到目标税率
                    if target_tax and target_tax in merged:
                        for col, src_tax in COL_TO_TAX.items():
                            known_price = prices.get(col)
                            if known_price is not None and src_tax in merged:
                                net = net_from_inclusive(float(known_price), merged[src_tax])
                                calc = inclusive_from_net(net, merged[target_tax])
                                return calc, f"calc_from_{src_tax}"

                return None, "unavailable"

            def pick_quote_row(fid: int, cat_id: int) -> Optional[Dict[str, Optional[float]]]:
                for cn in cat_id_to_names.get(cat_id, []):
                    row = raw_price_map.get((fid, cn), {})
                    if row:
                        return row
                return None

            # 组合结果；总运费 = 每吨运费（元/吨）× 吨数
            t = float(tons)
            result: List[Dict[str, Any]] = []
            for (wid, fid), (wname, fname, freight) in freight_map.items():
                for cid in category_ids:
                    cat_name = cat_map.get(cid)
                    if cat_name is None:
                        continue
                    price, source = resolve_price(fid, cid)
                    merged = merge_factory_rates(tax_rate_map.get(fid, {}))
                    if price is not None and target_tax and target_tax in merged:
                        p_net = round(
                            net_from_inclusive(float(price), merged[target_tax]), 2
                        )
                    elif price is not None:
                        p_net = float(price)
                    else:
                        p_net = None

                    fr = float(freight) if freight is not None else 0.0
                    freight_cost_total = round(fr * t, 2)

                    quote_amount: Optional[float] = (
                        round(float(p_net) * t, 2) if p_net is not None else None
                    )
                    profit = (
                        round(quote_amount - freight_cost_total, 2)
                        if quote_amount is not None
                        else round(-freight_cost_total, 2)
                    )
                    p = float(p_net) if p_net is not None else 0.0

                    qrow = pick_quote_row(fid, cid)
                    breakdown = (
                        derive_net_and_vat_from_quote_row(qrow, merged) if qrow else None
                    )
                    if breakdown:
                        base_net, p1_vat, p3_vat, _p13 = breakdown
                        profit_base = round(base_net * t - freight_cost_total, 2)
                        profit_3 = round(p3_vat * t - freight_cost_total, 2)
                    else:
                        base_net = None
                        p1_vat = None
                        p3_vat = None
                        profit_base = None
                        profit_3 = None

                    optimal_profits: Dict[str, Optional[float]] = {}
                    for b in bases:
                        u = _unit_for_optimal_price_basis(b, breakdown, qrow)
                        optimal_profits[b] = (
                            round(u * t - freight_cost_total, 2) if u is not None else None
                        )

                    rec: Dict[str, Any] = {
                        "仓库id": wid,
                        "冶炼厂id": fid,
                        "品类id": cid,
                        "仓库": wname,
                        "冶炼厂": fname,
                        "品类": cat_name,
                        "price_type": price_type_name,
                        "吨数": t,
                        "运费计价方式": "per_ton",
                        "运费": fr,
                        "总运费": freight_cost_total,
                        "报价": p_net if source != "unavailable" else None,
                        "报价金额": quote_amount,
                        "报价来源": source,
                        "基准价": base_net,
                        "含1%税价": p1_vat,
                        "含3%税价": p3_vat,
                        "利润": profit,
                        "利润_基准": profit_base,
                        "利润_含3%": profit_3,
                        "最优价各口径利润": optimal_profits,
                    }
                    result.append(rec)

            result.sort(
                key=lambda r: (
                    r["最优价各口径利润"][sort_basis]
                    if r["最优价各口径利润"].get(sort_basis) is not None
                    else float("-inf")
                ),
                reverse=True,
            )

            # 按冶炼厂汇总；排行按「最优价排序口径」对应利润合计从高到低
            agg: Dict[int, Dict[str, Any]] = {}
            for row in result:
                sfid = int(row["冶炼厂id"])
                if sfid not in agg:
                    agg[sfid] = {
                        "冶炼厂id": sfid,
                        "冶炼厂": row["冶炼厂"],
                        "利润": 0.0,
                        "利润_含3%合计": 0.0,
                        "利润_基准合计": 0.0,
                        "最优价口径合计": {b: 0.0 for b in bases},
                    }
                agg[sfid]["利润"] += float(row["利润"])
                if row["利润_含3%"] is not None:
                    agg[sfid]["利润_含3%合计"] += float(row["利润_含3%"])
                if row["利润_基准"] is not None:
                    agg[sfid]["利润_基准合计"] += float(row["利润_基准"])
                op = row["最优价各口径利润"]
                for b in bases:
                    pv = op.get(b)
                    if pv is not None:
                        agg[sfid]["最优价口径合计"][b] += float(pv)

            ranking = sorted(
                (
                    {
                        **v,
                        "利润": round(v["利润"], 2),
                        "利润_含3%合计": round(v["利润_含3%合计"], 2),
                        "利润_基准合计": round(v["利润_基准合计"], 2),
                        "最优价口径合计": {
                            b: round(v["最优价口径合计"][b], 2) for b in bases
                        },
                    }
                    for v in agg.values()
                ),
                key=lambda x: x["最优价口径合计"][sort_basis],
                reverse=True,
            )
            return {
                "明细": result,
                "冶炼厂利润排行": ranking,
                "最优价排序口径": sort_basis,
            }

        except Exception as e:
            logger.error(f"获取比价表失败: {e}")
            raise

    # ==================== 税率表 CRUD ====================

    def get_tax_rates(self, factory_ids: Optional[List[int]] = None) -> List[Dict[str, Any]]:
        """获取税率表，可按冶炼厂过滤"""
        try:
            with get_conn() as conn:
                with conn.cursor() as cur:
                    if factory_ids:
                        ph = ",".join(["%s"] * len(factory_ids))
                        cur.execute(
                            f"SELECT ftr.id, ftr.factory_id, df.name AS factory_name, "
                            f"ftr.tax_type, ftr.tax_rate "
                            f"FROM factory_tax_rates ftr "
                            f"JOIN dict_factories df ON ftr.factory_id = df.id "
                            f"WHERE ftr.factory_id IN ({ph}) "
                            f"ORDER BY ftr.factory_id, ftr.tax_type",
                            tuple(factory_ids),
                        )
                    else:
                        cur.execute(
                            "SELECT ftr.id, ftr.factory_id, df.name AS factory_name, "
                            "ftr.tax_type, ftr.tax_rate "
                            "FROM factory_tax_rates ftr "
                            "JOIN dict_factories df ON ftr.factory_id = df.id "
                            "ORDER BY ftr.factory_id, ftr.tax_type"
                        )
                    cols = [d[0] for d in cur.description]
                    return [dict(zip(cols, row)) for row in cur.fetchall()]
        except Exception as e:
            logger.error(f"获取税率表失败: {e}")
            raise

    def upsert_tax_rates(self, items: List[Dict[str, Any]]) -> Dict[str, Any]:
        """批量设置税率（存在则更新，不存在则插入）"""
        from app.models.tl import VALID_TAX_TYPES
        for item in items:
            if item["tax_type"] not in VALID_TAX_TYPES:
                raise ValueError(f"不支持的 tax_type: {item['tax_type']}，有效值：{VALID_TAX_TYPES}")
            if not (0 <= item["tax_rate"] <= 1):
                raise ValueError(f"tax_rate 必须在 0~1 之间，收到：{item['tax_rate']}")
        try:
            with get_conn() as conn:
                with conn.cursor() as cur:
                    for item in items:
                        # 验证冶炼厂是否存在
                        cur.execute("SELECT id FROM dict_factories WHERE id = %s", (item["factory_id"],))
                        if not cur.fetchone():
                            raise ValueError(f"冶炼厂 ID {item['factory_id']} 不存在")

                        cur.execute(
                            "INSERT INTO factory_tax_rates (factory_id, tax_type, tax_rate) "
                            "VALUES (%s, %s, %s) "
                            "ON DUPLICATE KEY UPDATE tax_rate = VALUES(tax_rate), "
                            "updated_at = CURRENT_TIMESTAMP",
                            (item["factory_id"], item["tax_type"], item["tax_rate"]),
                        )
            return {"code": 200, "msg": f"已保存 {len(items)} 条税率记录"}
        except ValueError:
            raise
        except Exception as e:
            logger.error(f"设置税率失败: {e}")
            raise

    def delete_tax_rate(self, factory_id: int, tax_type: str) -> Dict[str, Any]:
        """删除某冶炼厂的某税率记录"""
        try:
            with get_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        "DELETE FROM factory_tax_rates WHERE factory_id = %s AND tax_type = %s",
                        (factory_id, tax_type),
                    )
                    if cur.rowcount == 0:
                        raise ValueError(f"未找到 factory_id={factory_id}, tax_type={tax_type} 的记录")
            return {"code": 200, "msg": "删除成功"}
        except ValueError:
            raise
        except Exception as e:
            logger.error(f"删除税率失败: {e}")
            raise

    # ==================== 接口5：上传价格表（OCR解析） ====================

    def _match_factory(
        self, ocr_name: str, factory_list: List[Tuple[int, str]]
    ) -> Optional[int]:
        """将 OCR 识别出的工厂名匹配到 dict_factories 中的冶炼厂，返回 factory_id"""
        if not ocr_name or ocr_name == "未知工厂":
            return None
        for fid, fname in factory_list:
            # 双向包含匹配
            if fname in ocr_name or ocr_name in fname:
                return fid
        return None

    def _match_category(
        self, ocr_cat: str, category_list: List[Tuple[int, int, str]]
    ) -> Optional[Tuple[int, int]]:
        """将 OCR 识别出的品类名匹配到 dict_categories，返回 (category_id, row_id)"""
        if not ocr_cat:
            return None
        for row_id, cat_id, cname in category_list:
            if cname in ocr_cat or ocr_cat in cname:
                return (cat_id, row_id)
        return None

    def upload_price_table(self, files: List[Any]) -> Dict[str, Any]:
        saved_paths: List[Tuple[str, str, str]] = []
        try:
            # 1. 保存图片到磁盘
            for upload_file in files:
                content = upload_file.file.read()
                md5 = hashlib.md5(content).hexdigest()
                suffix = Path(upload_file.filename).suffix or ".jpg"
                filename = f"{uuid.uuid4().hex}{suffix}"
                save_path = PRICE_TABLE_UPLOAD_DIR / filename

                with open(save_path, "wb") as f:
                    f.write(content)
                saved_paths.append((str(save_path), md5, upload_file.filename))

            # 2. VLM识别
            from app import config as app_config
            if not app_config.VLM_API_KEY:
                raise ValueError("未配置 VLM_API_KEY，请在环境变量中设置 VLM_API_KEY")
            vlm_config = VLMConfig(
                api_key=app_config.VLM_API_KEY,
                base_url=app_config.VLM_BASE_URL,
                model=app_config.VLM_MODEL,
                max_tokens=app_config.VLM_MAX_TOKENS,
                image_max_edge=app_config.VLM_IMAGE_MAX_EDGE,
                jpeg_quality=app_config.VLM_JPEG_QUALITY,
                request_timeout=app_config.VLM_REQUEST_TIMEOUT,
                save_individual=False,
            )

            details = []
            with QwenVLFullExtractor(vlm_config) as extractor:
                for image_path, md5, orig_name in saved_paths:
                    result = extractor.recognize(image_path, save_output=False)

                    if not result.success:
                        details.append({
                            "image": orig_name,
                            "success": False,
                            "error": result.error_message,
                        })
                        continue

                    # 3. 构建 full_data（VlmFullData格式，供前端保留并回传）
                    full_data = {
                        "image_path": result.image_path,
                        "file_name": result.file_name,
                        "source_image": orig_name,
                        "company_name": result.company_name,
                        "doc_title": result.doc_title,
                        "subtitle": result.subtitle,
                        "quote_date": result.quote_date,
                        "execution_date": result.execution_date,
                        "valid_period": result.valid_period,
                        "price_unit": result.price_unit,
                        "headers": result.headers,
                        "rows": [row.model_dump() for row in result.rows],
                        "policies": result.policies,
                        "footer_notes": result.footer_notes,
                        "footer_notes_raw": result.footer_notes_raw,
                        "brand_specifications": result.brand_specifications,
                        "raw_full_text": result.raw_full_text,
                        "elapsed_time": result.elapsed_time,
                    }

                    # 4. 映射为前端可编辑的 items（ConfirmPriceTableItem格式）
                    items = self._map_vlm_to_confirm_items(result)

                    details.append({
                        "image": orig_name,
                        "success": True,
                        "full_data": full_data,
                        "items": items,
                    })

            return {"code": 200, "data": {"details": details}}

        except Exception as e:
            logger.error(f"上传价格表失败: {e}")
            for path, _, _ in saved_paths:
                try:
                    os.remove(path)
                except OSError:
                    pass
            raise

    def _map_vlm_to_confirm_items(self, result) -> List[Dict[str, Any]]:
        """将 VLM 结果映射为确认条目：图上可能是基准价或含税价（多列/备注）。此处仅带出 OCR 显式列与预览用不含税/反算（默认税率占位）；确认写入时用冶炼厂系统税率做「基准↔含税」双向统一。"""
        items = []
        factory_name = result.company_name or ""
        defaults = merge_factory_rates(None)
        for row in result.rows:
            price_normal = row.price_normal_invoice
            price_reverse = row.price_reverse_invoice
            src: Dict[str, str] = {}

            if row.exclusive_net is not None:
                net_f = round(float(row.exclusive_net), 2)
                basis = getattr(row, "price_basis", None) or parse_price_basis_from_remark(
                    row.remark
                )
                fp1 = float(row.price_1pct_vat) if row.price_1pct_vat is not None else None
                fp3 = float(row.price_3pct_vat) if row.price_3pct_vat is not None else None
                fp13 = float(row.price_13pct_vat) if row.price_13pct_vat is not None else None
                src["unit_price"] = SOURCE_ORIGINAL
                if fp1 is not None:
                    src["price_1pct_vat"] = SOURCE_ORIGINAL
                if fp3 is not None:
                    src["price_3pct_vat"] = SOURCE_ORIGINAL
                if fp13 is not None:
                    src["price_13pct_vat"] = SOURCE_ORIGINAL
            else:
                basis = parse_price_basis_from_remark(row.remark)
                pg = row.price_general
                if pg is not None:
                    net_f, _, _, _ = derive_vat_prices_from_stated_price(
                        float(pg), basis, None
                    )
                    net_f = round(net_f, 2)
                    fp1 = float(row.price_1pct_vat) if row.price_1pct_vat is not None else None
                    fp3 = float(row.price_3pct_vat) if row.price_3pct_vat is not None else None
                    fp13 = float(row.price_13pct_vat) if row.price_13pct_vat is not None else None
                    src["unit_price"] = SOURCE_DERIVED
                    if fp1 is not None:
                        src["price_1pct_vat"] = SOURCE_ORIGINAL
                    if fp3 is not None:
                        src["price_3pct_vat"] = SOURCE_ORIGINAL
                    if fp13 is not None:
                        src["price_13pct_vat"] = SOURCE_ORIGINAL
                elif row.price_3pct_vat is not None:
                    net_f = round(
                        net_from_inclusive(float(row.price_3pct_vat), defaults["3pct"]), 2
                    )
                    fp1 = float(row.price_1pct_vat) if row.price_1pct_vat is not None else None
                    fp3 = float(row.price_3pct_vat)
                    fp13 = float(row.price_13pct_vat) if row.price_13pct_vat is not None else None
                    src["unit_price"] = SOURCE_DERIVED
                    src["price_3pct_vat"] = SOURCE_ORIGINAL
                    if fp1 is not None:
                        src["price_1pct_vat"] = SOURCE_ORIGINAL
                    if fp13 is not None:
                        src["price_13pct_vat"] = SOURCE_ORIGINAL
                elif row.price_13pct_vat is not None:
                    net_f = round(
                        net_from_inclusive(float(row.price_13pct_vat), defaults["13pct"]), 2
                    )
                    fp1 = float(row.price_1pct_vat) if row.price_1pct_vat is not None else None
                    fp3 = float(row.price_3pct_vat) if row.price_3pct_vat is not None else None
                    fp13 = float(row.price_13pct_vat)
                    src["unit_price"] = SOURCE_DERIVED
                    src["price_13pct_vat"] = SOURCE_ORIGINAL
                    if fp1 is not None:
                        src["price_1pct_vat"] = SOURCE_ORIGINAL
                    if fp3 is not None:
                        src["price_3pct_vat"] = SOURCE_ORIGINAL
                elif row.price_1pct_vat is not None:
                    net_f = round(
                        net_from_inclusive(float(row.price_1pct_vat), defaults["1pct"]), 2
                    )
                    fp1 = float(row.price_1pct_vat)
                    fp3 = float(row.price_3pct_vat) if row.price_3pct_vat is not None else None
                    fp13 = float(row.price_13pct_vat) if row.price_13pct_vat is not None else None
                    src["unit_price"] = SOURCE_DERIVED
                    src["price_1pct_vat"] = SOURCE_ORIGINAL
                    if fp3 is not None:
                        src["price_3pct_vat"] = SOURCE_ORIGINAL
                    if fp13 is not None:
                        src["price_13pct_vat"] = SOURCE_ORIGINAL
                else:
                    net_f, fp1, fp3, fp13 = None, None, None, None

            if price_normal is not None:
                src["price_normal_invoice"] = SOURCE_ORIGINAL
            if price_reverse is not None:
                src["price_reverse_invoice"] = SOURCE_ORIGINAL

            items.append({
                "冶炼厂名": factory_name,
                "冶炼厂id": None,
                "品类名": row.category,
                "品类id": None,
                "价格": net_f,
                "价格口径": basis,
                "备注": row.remark or None,
                "价格_1pct增值税": fp1,
                "价格_3pct增值税": fp3,
                "价格_13pct增值税": fp13,
                "普通发票价格": float(price_normal) if price_normal is not None else None,
                "反向发票价格": float(price_reverse) if price_reverse is not None else None,
                "价格字段来源": src,
            })
        return items

    # ==================== 接口5b：确认价格表写入数据库 ====================

    def confirm_price_table(
        self,
        quote_date_str: str,
        items: List[Dict[str, Any]],
        full_data: Optional[Dict[str, Any]] = None,
    ) -> Dict[str, Any]:
        if not items:
            raise ValueError("报价数据不能为空")

        try:
            quote_dt = date.fromisoformat(quote_date_str)
        except (ValueError, TypeError):
            raise ValueError(f"日期格式不正确: {quote_date_str}，应为 YYYY-MM-DD")

        try:
            with get_conn() as conn:
                with conn.cursor() as cur:
                    inserted, updated = 0, 0

                    for item in items:
                        # 1. 冶炼厂不存在则新建
                        if item.get("冶炼厂id") is None:
                            factory_name = item["冶炼厂名"]
                            cur.execute(
                                "SELECT id FROM dict_factories WHERE name = %s",
                                (factory_name,),
                            )
                            row = cur.fetchone()
                            if row:
                                item["冶炼厂id"] = row[0]
                            else:
                                cur.execute(
                                    "INSERT INTO dict_factories (name, is_active) "
                                    "VALUES (%s, 1)",
                                    (factory_name,),
                                )
                                item["冶炼厂id"] = cur.lastrowid

                        # 2. 品类不存在则新建到 dict_categories
                        cat_name = item["品类名"]
                        cur.execute(
                            "SELECT category_id FROM dict_categories WHERE name = %s AND is_active = 1",
                            (cat_name,),
                        )
                        row = cur.fetchone()
                        if not row:
                            # 新建品类，分配新的 category_id
                            cur.execute("SELECT COALESCE(MAX(category_id), 0) + 1 FROM dict_categories")
                            new_cat_id = cur.fetchone()[0]
                            cur.execute(
                                "INSERT INTO dict_categories "
                                "(category_id, name, is_main, is_active) "
                                "VALUES (%s, %s, 1, 1)",
                                (new_cat_id, cat_name),
                            )

                    # 3. 存储全量元数据（如果有 full_data）
                    metadata_id = None
                    if full_data:
                        # 取第一条 item 的冶炼厂id作为元数据的 factory_id
                        factory_id_for_meta = items[0].get("冶炼厂id") if items else None
                        if factory_id_for_meta:
                            cur.execute(
                                """
                                INSERT INTO quote_table_metadata
                                (factory_id, quote_date, execution_date, doc_title, subtitle,
                                 valid_period, price_unit, headers, footer_notes, footer_notes_raw,
                                 brand_specifications, policies, raw_full_text, source_image)
                                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                                ON DUPLICATE KEY UPDATE
                                    execution_date = VALUES(execution_date),
                                    doc_title = VALUES(doc_title),
                                    subtitle = VALUES(subtitle),
                                    valid_period = VALUES(valid_period),
                                    price_unit = VALUES(price_unit),
                                    headers = VALUES(headers),
                                    footer_notes = VALUES(footer_notes),
                                    footer_notes_raw = VALUES(footer_notes_raw),
                                    brand_specifications = VALUES(brand_specifications),
                                    policies = VALUES(policies),
                                    raw_full_text = VALUES(raw_full_text),
                                    source_image = VALUES(source_image),
                                    updated_at = CURRENT_TIMESTAMP
                                """,
                                (
                                    factory_id_for_meta,
                                    quote_dt,
                                    full_data.get("execution_date", ""),
                                    full_data.get("doc_title", ""),
                                    full_data.get("subtitle", ""),
                                    full_data.get("valid_period", ""),
                                    full_data.get("price_unit", "元/吨"),
                                    json.dumps(full_data.get("headers", []), ensure_ascii=False),
                                    json.dumps(full_data.get("footer_notes", []), ensure_ascii=False),
                                    full_data.get("footer_notes_raw", ""),
                                    full_data.get("brand_specifications", ""),
                                    json.dumps(full_data.get("policies", {}), ensure_ascii=False),
                                    full_data.get("raw_full_text", ""),
                                    full_data.get("source_image", full_data.get("file_name", "")),
                                ),
                            )
                            # 取 metadata_id（INSERT 或 已存在的）
                            if cur.lastrowid:
                                metadata_id = cur.lastrowid
                            else:
                                cur.execute(
                                    "SELECT id FROM quote_table_metadata WHERE factory_id=%s AND quote_date=%s",
                                    (factory_id_for_meta, quote_dt),
                                )
                                row = cur.fetchone()
                                metadata_id = row[0] if row else None

                    # 3b. 按冶炼厂 factory_tax_rates（与默认合并）统一计算「价格」与含1%/3%/13%价（覆盖上传预览推算）
                    factory_ids = list({item["冶炼厂id"] for item in items})
                    tax_by_fid: Dict[int, Dict[str, float]] = {}
                    if factory_ids:
                        fph = ",".join(["%s"] * len(factory_ids))
                        cur.execute(
                            f"SELECT factory_id, tax_type, tax_rate FROM factory_tax_rates "
                            f"WHERE factory_id IN ({fph})",
                            tuple(factory_ids),
                        )
                        for fid, ttype, tr in cur.fetchall():
                            tax_by_fid.setdefault(int(fid), {})[str(ttype)] = float(tr)
                    snapshots = [{k: it.get(k) for k in API_KEY_TO_DB} for it in items]

                    applied_factory_tax: List[bool] = []
                    for item in items:
                        applied_factory_tax.append(
                            _apply_factory_tax_rates_to_quote_item(item, tax_by_fid)
                        )

                    final_sources_list: List[Dict[str, str]] = []
                    for item, snap, tax_applied in zip(items, snapshots, applied_factory_tax):
                        client_src = normalize_client_sources(item.get("价格字段来源"))
                        merged_src = merge_sources_after_fill(item, snap, client_src)
                        if tax_applied:
                            merged_src["price_1pct_vat"] = SOURCE_DERIVED
                            merged_src["price_3pct_vat"] = SOURCE_DERIVED
                            merged_src["price_13pct_vat"] = SOURCE_DERIVED
                            merged_src["unit_price"] = (
                                SOURCE_ORIGINAL
                                if snap.get("价格") is not None
                                else SOURCE_DERIVED
                            )
                        final_sources_list.append(merged_src)

                    # 4. 写入明细，相同(日期+冶炼厂+品类名)则更新价格
                    written_sources: List[Dict[str, Any]] = []
                    for item, final_src in zip(items, final_sources_list):
                        src_json = json.dumps(final_src, ensure_ascii=False) if final_src else None
                        cur.execute(
                            """
                            INSERT INTO quote_details
                            (quote_date, factory_id, category_name, metadata_id,
                             unit_price, price_1pct_vat, price_3pct_vat, price_13pct_vat,
                             price_normal_invoice, price_reverse_invoice, price_field_sources)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                            ON DUPLICATE KEY UPDATE
                                metadata_id = VALUES(metadata_id),
                                unit_price = VALUES(unit_price),
                                price_1pct_vat = VALUES(price_1pct_vat),
                                price_3pct_vat = VALUES(price_3pct_vat),
                                price_13pct_vat = VALUES(price_13pct_vat),
                                price_normal_invoice = VALUES(price_normal_invoice),
                                price_reverse_invoice = VALUES(price_reverse_invoice),
                                price_field_sources = VALUES(price_field_sources),
                                updated_at = CURRENT_TIMESTAMP
                            """,
                            (
                                quote_dt,
                                item["冶炼厂id"],
                                item["品类名"],
                                metadata_id,
                                item.get("价格"),
                                item.get("价格_1pct增值税"),
                                item.get("价格_3pct增值税"),
                                item.get("价格_13pct增值税"),
                                item.get("普通发票价格"),
                                item.get("反向发票价格"),
                                src_json,
                            ),
                        )
                        if cur.rowcount == 1:
                            inserted += 1
                        else:
                            updated += 1
                        written_sources.append(
                            {
                                "冶炼厂id": item["冶炼厂id"],
                                "品类名": item["品类名"],
                                "价格字段来源": final_src,
                            }
                        )

            return {
                "code": 200,
                "msg": f"写入成功：新增 {inserted} 条，更新 {updated} 条",
                "明细价格来源": written_sources,
            }

        except ValueError:
            raise
        except Exception as e:
            logger.error(f"确认价格表写入失败: {e}")
            raise

    # ==================== 接口6：上传运费 ====================

    def upload_freight(self, freight_list: List[Dict[str, Any]]) -> Dict[str, Any]:
        try:
            with get_conn() as conn:
                with conn.cursor() as cur:
                    today = date.today().isoformat()
                    for item in freight_list:
                        warehouse_name = item["仓库"]
                        smelter_name = item["冶炼厂"]
                        freight = item["运费"]

                        cur.execute(
                            "SELECT id FROM dict_warehouses WHERE name = %s AND is_active = 1",
                            (warehouse_name,),
                        )
                        wh_row = cur.fetchone()
                        if not wh_row:
                            raise ValueError(f"仓库 '{warehouse_name}' 不存在或未启用")

                        cur.execute(
                            "SELECT id FROM dict_factories WHERE name = %s AND is_active = 1",
                            (smelter_name,),
                        )
                        sm_row = cur.fetchone()
                        if not sm_row:
                            raise ValueError(f"冶炼厂 '{smelter_name}' 不存在或未启用")

                        cur.execute(
                            "INSERT INTO freight_rates "
                            "(factory_id, warehouse_id, price_per_ton, effective_date) "
                            "VALUES (%s, %s, %s, %s) "
                            "ON DUPLICATE KEY UPDATE "
                            "price_per_ton = VALUES(price_per_ton), "
                            "updated_at = CURRENT_TIMESTAMP",
                            (sm_row[0], wh_row[0], freight, today),
                        )
            return {"code": 200, "msg": "运费数据已存入数据库"}

        except ValueError:
            raise
        except Exception as e:
            logger.error(f"上传运费失败: {e}")
            raise

    def build_freight_template_excel(self, warehouse_ids: List[int]) -> bytes:
        """首列：所选库房名称（按传入 id 顺序）；表头：库房 + 全部启用冶炼厂；数据格留空。"""
        if not warehouse_ids:
            raise ValueError("库房id列表不能为空")
        seen: set[int] = set()
        ordered_ids: List[int] = []
        for wid in warehouse_ids:
            if wid in seen:
                continue
            seen.add(wid)
            ordered_ids.append(int(wid))
        try:
            from openpyxl import Workbook

            with get_conn() as conn:
                with conn.cursor() as cur:
                    wh_ph = ",".join(["%s"] * len(ordered_ids))
                    cur.execute(
                        f"SELECT id, name FROM dict_warehouses "
                        f"WHERE id IN ({wh_ph}) AND is_active = 1",
                        tuple(ordered_ids),
                    )
                    wh_map: Dict[int, str] = {int(r[0]): str(r[1]) for r in cur.fetchall()}
                    missing = [i for i in ordered_ids if i not in wh_map]
                    if missing:
                        raise ValueError(f"以下库房不存在或未启用: {missing}")

                    cur.execute(
                        "SELECT name FROM dict_factories WHERE is_active = 1 ORDER BY id"
                    )
                    smelter_names = [str(r[0]) for r in cur.fetchall()]
                    if not smelter_names:
                        raise ValueError("没有可用的冶炼厂，请先在冶炼厂字典中维护")

            wb = Workbook()
            ws = wb.active
            ws.title = "运费配置"
            header = ["库房"] + smelter_names
            ws.append(header)
            for wid in ordered_ids:
                ws.append([wh_map[wid]] + [None] * len(smelter_names))
            buf = io.BytesIO()
            wb.save(buf)
            return buf.getvalue()
        except ValueError:
            raise
        except Exception as e:
            logger.error(f"生成运费模板 Excel 失败: {e}")
            raise

    def import_freight_excel(self, content: bytes) -> Dict[str, Any]:
        """
        解析运费矩阵：第 1 行表头为「库房」+ 冶炼厂名称；自第 2 行起首列为库房名，对应列填运费。
        与 upload_freight 相同写入 freight_rates（当日生效）；空单元格跳过。
        表头或首列出现库中不存在的冶炼厂、库房名称时，自动写入 dict_factories / dict_warehouses
       （与 add_smelter / add_warehouse 一致；若名称已存在但已停用则恢复启用）。
        """
        if not content:
            raise ValueError("文件内容为空")
        try:
            from openpyxl import load_workbook
        except ImportError as e:
            raise ValueError("服务端未安装 openpyxl，无法导入 Excel") from e

        def _coerce_freight(v: Any) -> Optional[float]:
            if v is None:
                return None
            if isinstance(v, bool):
                return None
            if isinstance(v, str):
                s = v.strip()
                if not s:
                    return None
                v = s
            try:
                x = float(v)
            except (TypeError, ValueError):
                return None
            if x < 0:
                raise ValueError("运费不能为负数")
            return round(x, 2)

        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        try:
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            header_row = next(rows_iter, None)
            if not header_row or not any(
                c is not None and str(c).strip() for c in header_row
            ):
                raise ValueError("无法读取表头（第 1 行）")

            headers = [str(c).strip() if c is not None else "" for c in header_row]
            if not headers:
                raise ValueError("表头为空")

            factory_by_col: Dict[int, Tuple[str, int]] = {}
            stats = {
                "new_wh": 0,
                "new_fa": 0,
                "re_wh": 0,
                "re_fa": 0,
            }

            with get_conn() as conn:
                with conn.cursor() as cur:

                    def _ensure_warehouse_id(name: str) -> int:
                        cur.execute(
                            "SELECT id, is_active FROM dict_warehouses WHERE name = %s",
                            (name,),
                        )
                        row = cur.fetchone()
                        if row:
                            wid, act = int(row[0]), row[1]
                            if act != 1:
                                cur.execute(
                                    "UPDATE dict_warehouses SET is_active = 1 WHERE id = %s",
                                    (wid,),
                                )
                                stats["re_wh"] += 1
                            return wid
                        cur.execute(
                            "INSERT INTO dict_warehouses (name, is_active) VALUES (%s, 1)",
                            (name,),
                        )
                        stats["new_wh"] += 1
                        return int(cur.lastrowid)

                    def _ensure_factory_id(name: str) -> int:
                        cur.execute(
                            "SELECT id, is_active FROM dict_factories WHERE name = %s",
                            (name,),
                        )
                        row = cur.fetchone()
                        if row:
                            fid, act = int(row[0]), row[1]
                            if act != 1:
                                cur.execute(
                                    "UPDATE dict_factories SET is_active = 1 WHERE id = %s",
                                    (fid,),
                                )
                                stats["re_fa"] += 1
                            return fid
                        cur.execute(
                            "INSERT INTO dict_factories (name, is_active) VALUES (%s, 1)",
                            (name,),
                        )
                        stats["new_fa"] += 1
                        return int(cur.lastrowid)

                    for col_idx in range(1, len(headers)):
                        h = headers[col_idx]
                        if not h:
                            continue
                        fid = _ensure_factory_id(h)
                        factory_by_col[col_idx + 1] = (h, fid)

                    if not factory_by_col:
                        raise ValueError(
                            "表头从第 2 列起至少需要一列冶炼厂名称（表头不能为空）"
                        )

                    today = date.today().isoformat()
                    written = 0
                    skipped_rows = 0
                    skipped_cells = 0
                    errors: List[str] = []

                    for ridx, row in enumerate(rows_iter, start=2):
                        if not row:
                            skipped_rows += 1
                            continue
                        wh_cell = row[0] if len(row) > 0 else None
                        if wh_cell is None or (
                            isinstance(wh_cell, str) and not wh_cell.strip()
                        ):
                            skipped_rows += 1
                            continue
                        wh_name = str(wh_cell).strip()
                        try:
                            wid = _ensure_warehouse_id(wh_name)
                        except Exception as ex:
                            errors.append(f"第{ridx}行：库房「{wh_name}」未能写入字典：{ex}")
                            continue

                        for col_idx, (fname, fid) in factory_by_col.items():
                            if col_idx - 1 >= len(row):
                                continue
                            cell_v = row[col_idx - 1]
                            freight = _coerce_freight(cell_v)
                            if freight is None:
                                skipped_cells += 1
                                continue
                            cur.execute(
                                "INSERT INTO freight_rates "
                                "(factory_id, warehouse_id, price_per_ton, effective_date) "
                                "VALUES (%s, %s, %s, %s) "
                                "ON DUPLICATE KEY UPDATE "
                                "price_per_ton = VALUES(price_per_ton), "
                                "updated_at = CURRENT_TIMESTAMP",
                                (fid, wid, freight, today),
                            )
                            written += 1

            msg = f"已写入 {written} 条运费（生效日期 {today}）"
            extra = []
            if stats["new_wh"] or stats["re_wh"]:
                extra.append(f"库房新建 {stats['new_wh']}、恢复启用 {stats['re_wh']}")
            if stats["new_fa"] or stats["re_fa"]:
                extra.append(f"冶炼厂新建 {stats['new_fa']}、恢复启用 {stats['re_fa']}")
            if extra:
                msg += "；" + "；".join(extra)
            return {
                "code": 200,
                "msg": msg,
                "写入条数": written,
                "新建库房数": stats["new_wh"],
                "恢复启用库房数": stats["re_wh"],
                "新建冶炼厂数": stats["new_fa"],
                "恢复启用冶炼厂数": stats["re_fa"],
                "跳过空单元格数": skipped_cells,
                "跳过空行数": skipped_rows,
                "错误明细": errors if errors else None,
            }
        except ValueError:
            raise
        except Exception as e:
            logger.error(f"导入运费 Excel 失败: {e}")
            raise
        finally:
            wb.close()

    def _build_factory_latest_quote_catalog(
        self, factory_ids: List[int]
    ) -> Dict[int, List[Dict[str, Any]]]:
        """
        每个冶炼厂在系统品类下的「最新」报价：同一品类多别名时取 quote_date 最新的一条；无记录则各价为 null。
        与比价接口取价一致（按冶炼厂+品种名称维度的 MAX(quote_date)）。
        """
        if not factory_ids:
            return {}
        fac_ph = ",".join(["%s"] * len(factory_ids))
        try:
            with get_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        "SELECT category_id, name FROM dict_categories "
                        "WHERE is_active = 1 ORDER BY category_id, is_main DESC, row_id"
                    )
                    cat_id_to_names: Dict[int, List[str]] = {}
                    cat_id_main: Dict[int, str] = {}
                    for cid, name in cur.fetchall():
                        cid = int(cid)
                        n = str(name).strip()
                        if not n:
                            continue
                        cat_id_to_names.setdefault(cid, []).append(n)
                        if cid not in cat_id_main:
                            cat_id_main[cid] = n

                    cur.execute(
                        f"""
                        SELECT qd.factory_id, qd.category_name, qd.quote_date,
                               qd.unit_price, qd.price_3pct_vat, qd.price_13pct_vat
                        FROM quote_details qd
                        JOIN (
                            SELECT factory_id, category_name, MAX(quote_date) AS mq
                            FROM quote_details
                            WHERE factory_id IN ({fac_ph})
                            GROUP BY factory_id, category_name
                        ) t ON qd.factory_id = t.factory_id
                           AND qd.category_name = t.category_name
                           AND qd.quote_date = t.mq
                        WHERE qd.factory_id IN ({fac_ph})
                        """,
                        tuple(factory_ids) + tuple(factory_ids),
                    )
                    # (fid, name) -> (quote_date, unit, p3, p13)
                    latest_by_pair: Dict[Tuple[int, str], Tuple[Any, Any, Any, Any]] = {}
                    for fid, cname, qd_d, up, p3, p13 in cur.fetchall():
                        latest_by_pair[(int(fid), str(cname).strip())] = (
                            qd_d,
                            up,
                            p3,
                            p13,
                        )

            out: Dict[int, List[Dict[str, Any]]] = {
                int(fid): [] for fid in factory_ids
            }
            sorted_cids = sorted(cat_id_to_names.keys())
            for fid in factory_ids:
                for cid in sorted_cids:
                    display = cat_id_main.get(cid, cat_id_to_names[cid][0])
                    best: Optional[Tuple[Any, Any, Any, Any]] = None
                    best_d: Optional[date] = None
                    for alias in cat_id_to_names[cid]:
                        key = (fid, alias)
                        if key not in latest_by_pair:
                            continue
                        qd_d, up, p3, p13 = latest_by_pair[key]
                        cmp_d = qd_d
                        if isinstance(cmp_d, datetime):
                            cmp_d = cmp_d.date()
                        if best_d is None or (
                            isinstance(cmp_d, date) and cmp_d > best_d
                        ):
                            best_d = cmp_d if isinstance(cmp_d, date) else None
                            best = (qd_d, up, p3, p13)
                    if best is None:
                        out[fid].append(
                            {
                                "品类id": cid,
                                "品种": display,
                                "报价日期": None,
                                "普通价": None,
                                "3%含税价": None,
                                "13%含税价": None,
                            }
                        )
                    else:
                        qd_d, up, p3, p13 = best
                        out[fid].append(
                            {
                                "品类id": cid,
                                "品种": display,
                                "报价日期": _cell_json(qd_d),
                                "普通价": _cell_json(up),
                                "3%含税价": _cell_json(p3),
                                "13%含税价": _cell_json(p13),
                            }
                        )
            return out
        except Exception as e:
            logger.error(f"构建冶炼厂最新报价目录失败: {e}")
            raise

    # ==================== 接口6b：运费列表 ====================

    def get_freight_list(
        self,
        warehouse_id: Optional[int] = None,
        factory_id: Optional[int] = None,
        date_from: Optional[str] = None,
        date_to: Optional[str] = None,
        page: int = 1,
        page_size: int = 50,
        include_latest_quotes: bool = False,
    ) -> Dict[str, Any]:
        if page < 1:
            raise ValueError("page 必须 >= 1")
        page_size = min(max(page_size, 1), 500)
        d_from: Optional[date] = None
        d_to: Optional[date] = None
        if date_from:
            try:
                d_from = date.fromisoformat(date_from)
            except (ValueError, TypeError):
                raise ValueError(f"date_from 格式不正确: {date_from}，应为 YYYY-MM-DD")
        if date_to:
            try:
                d_to = date.fromisoformat(date_to)
            except (ValueError, TypeError):
                raise ValueError(f"date_to 格式不正确: {date_to}，应为 YYYY-MM-DD")
        if d_from and d_to and d_from > d_to:
            raise ValueError("date_from 不能晚于 date_to")

        conditions: List[str] = ["1=1"]
        params: List[Any] = []
        if warehouse_id is not None:
            conditions.append("fr.warehouse_id = %s")
            params.append(warehouse_id)
        if factory_id is not None:
            conditions.append("fr.factory_id = %s")
            params.append(factory_id)
        if d_from is not None:
            conditions.append("fr.effective_date >= %s")
            params.append(d_from)
        if d_to is not None:
            conditions.append("fr.effective_date <= %s")
            params.append(d_to)
        where_sql = " AND ".join(conditions)
        offset = (page - 1) * page_size

        base_from = (
            "FROM freight_rates fr "
            "JOIN dict_warehouses dw ON fr.warehouse_id = dw.id "
            "JOIN dict_factories df ON fr.factory_id = df.id "
            f"WHERE {where_sql}"
        )
        try:
            with get_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute(f"SELECT COUNT(*) {base_from}", tuple(params))
                    total = cur.fetchone()[0]

                    cur.execute(
                        f"""
                        SELECT fr.id,
                               fr.warehouse_id AS `仓库id`,
                               dw.name AS `仓库名`,
                               fr.factory_id AS `冶炼厂id`,
                               df.name AS `冶炼厂`,
                               fr.price_per_ton AS `运费`,
                               fr.effective_date AS `生效日期`,
                               fr.created_at AS `创建时间`,
                               fr.updated_at AS `更新时间`
                        {base_from}
                        ORDER BY fr.effective_date DESC, fr.id DESC
                        LIMIT %s OFFSET %s
                        """,
                        tuple(params) + (page_size, offset),
                    )
                    cols = [d[0] for d in cur.description]
                    rows = [
                        {c: _cell_json(v) for c, v in zip(cols, r)}
                        for r in cur.fetchall()
                    ]
            data: Dict[str, Any] = {"total": total, "list": rows}
            if include_latest_quotes and rows:
                fac_ids = sorted({int(r["冶炼厂id"]) for r in rows})
                data["冶炼厂各品种最新报价"] = self._build_factory_latest_quote_catalog(
                    fac_ids
                )
            return {"code": 200, "data": data}
        except ValueError:
            raise
        except Exception as e:
            logger.error(f"获取运费列表失败: {e}")
            raise

    # ==================== 接口6c：编辑运费 ====================

    def update_freight(
        self,
        freight_id: int,
        price_per_ton: float,
        effective_date_str: Optional[str] = None,
    ) -> Dict[str, Any]:
        """按主键更新运费单价；可选修改生效日期（须满足 uk_factory_warehouse_date）。"""
        if freight_id < 1:
            raise ValueError("运费id 无效")
        new_ed: Optional[date] = None
        if effective_date_str is not None and str(effective_date_str).strip() != "":
            try:
                new_ed = date.fromisoformat(str(effective_date_str).strip())
            except (ValueError, TypeError):
                raise ValueError(f"生效日期格式不正确: {effective_date_str}，应为 YYYY-MM-DD")

        try:
            with get_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        "SELECT factory_id, warehouse_id, effective_date "
                        "FROM freight_rates WHERE id = %s",
                        (freight_id,),
                    )
                    row = cur.fetchone()
                    if not row:
                        raise ValueError(f"运费记录不存在: id={freight_id}")
                    factory_id, warehouse_id, current_ed = int(row[0]), int(row[1]), row[2]
                    if isinstance(current_ed, datetime):
                        current_ed = current_ed.date()

                    target_ed = new_ed if new_ed is not None else current_ed

                    if new_ed is not None and new_ed != current_ed:
                        cur.execute(
                            "SELECT id FROM freight_rates "
                            "WHERE factory_id = %s AND warehouse_id = %s "
                            "AND effective_date = %s AND id <> %s",
                            (factory_id, warehouse_id, new_ed, freight_id),
                        )
                        if cur.fetchone():
                            raise ValueError(
                                "该仓库与冶炼厂在目标生效日期已存在其它运费记录，无法改为该日期"
                            )

                    cur.execute(
                        "UPDATE freight_rates SET price_per_ton = %s, effective_date = %s, "
                        "updated_at = CURRENT_TIMESTAMP WHERE id = %s",
                        (price_per_ton, target_ed, freight_id),
                    )
                    if cur.rowcount == 0:
                        raise ValueError(f"更新失败: id={freight_id}")

            return {"code": 200, "msg": "运费已更新"}
        except ValueError:
            raise
        except Exception as e:
            logger.error(f"更新运费失败: {e}")
            raise

    # ==================== 接口6d：删除运费 ====================

    def delete_freight(self, freight_id: int) -> Dict[str, Any]:
        """按主键物理删除 `freight_rates` 一条记录（与 6c 使用同一 id）。"""
        if freight_id < 1:
            raise ValueError("运费id 无效")
        try:
            with get_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute("DELETE FROM freight_rates WHERE id = %s", (freight_id,))
                    if cur.rowcount == 0:
                        raise ValueError(f"运费记录不存在: id={freight_id}")
            return {"code": 200, "msg": "运费已删除"}
        except ValueError:
            raise
        except Exception as e:
            logger.error(f"删除运费失败: {e}")
            raise

    def _prepare_quote_details_filter(
        self,
        factory_id: Optional[int],
        quote_date: Optional[str],
        date_from: Optional[str],
        date_to: Optional[str],
        category_name: Optional[str],
        category_exact: bool,
    ) -> Tuple[str, List[Any]]:
        """报价明细列表/导出共用的 WHERE 与参数（含日期与品种校验）。"""
        qd_exact: Optional[date] = None
        d_from: Optional[date] = None
        d_to: Optional[date] = None
        if quote_date:
            try:
                qd_exact = date.fromisoformat(quote_date)
            except (ValueError, TypeError):
                raise ValueError(f"quote_date 格式不正确: {quote_date}，应为 YYYY-MM-DD")
        if date_from:
            try:
                d_from = date.fromisoformat(date_from)
            except (ValueError, TypeError):
                raise ValueError(f"date_from 格式不正确: {date_from}，应为 YYYY-MM-DD")
        if date_to:
            try:
                d_to = date.fromisoformat(date_to)
            except (ValueError, TypeError):
                raise ValueError(f"date_to 格式不正确: {date_to}，应为 YYYY-MM-DD")
        if d_from and d_to and d_from > d_to:
            raise ValueError("date_from 不能晚于 date_to")

        conditions: List[str] = ["1=1"]
        params: List[Any] = []
        if factory_id is not None:
            conditions.append("qd.factory_id = %s")
            params.append(factory_id)
        if qd_exact is not None:
            conditions.append("qd.quote_date = %s")
            params.append(qd_exact)
        if d_from is not None:
            conditions.append("qd.quote_date >= %s")
            params.append(d_from)
        if d_to is not None:
            conditions.append("qd.quote_date <= %s")
            params.append(d_to)
        if category_name:
            if category_exact:
                conditions.append("qd.category_name = %s")
                params.append(category_name)
            else:
                conditions.append("qd.category_name LIKE %s")
                params.append(f"%{category_name}%")
        where_sql = " AND ".join(conditions)
        return where_sql, params

    # ==================== 接口5c：报价数据列表 ====================

    def get_quote_details_list(
        self,
        factory_id: Optional[int] = None,
        quote_date: Optional[str] = None,
        date_from: Optional[str] = None,
        date_to: Optional[str] = None,
        category_name: Optional[str] = None,
        category_exact: bool = False,
        page: int = 1,
        page_size: int = 50,
        response_format: str = "full",
    ) -> Dict[str, Any]:
        if response_format not in ("full", "table"):
            raise ValueError('response_format 仅支持 "full" 或 "table"')
        if page < 1:
            raise ValueError("page 必须 >= 1")
        page_size = min(max(page_size, 1), 500)
        where_sql, params = self._prepare_quote_details_filter(
            factory_id=factory_id,
            quote_date=quote_date,
            date_from=date_from,
            date_to=date_to,
            category_name=category_name,
            category_exact=category_exact,
        )
        offset = (page - 1) * page_size

        base_from = (
            "FROM quote_details qd "
            "JOIN dict_factories df ON qd.factory_id = df.id "
            f"WHERE {where_sql}"
        )
        try:
            with get_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute(f"SELECT COUNT(*) {base_from}", tuple(params))
                    total = cur.fetchone()[0]

                    cur.execute(
                        f"""
                        SELECT qd.id,
                               qd.quote_date AS `报价日期`,
                               qd.factory_id AS `冶炼厂id`,
                               df.name AS `冶炼厂`,
                               qd.category_name AS `品类名`,
                               qd.metadata_id,
                               qd.unit_price AS `普通价`,
                               qd.price_1pct_vat AS `价格_1pct增值税`,
                               qd.price_3pct_vat AS `价格_3pct增值税`,
                               qd.price_13pct_vat AS `价格_13pct增值税`,
                               qd.price_normal_invoice AS `普通发票价格`,
                               qd.price_reverse_invoice AS `反向发票价格`,
                               qd.price_field_sources AS `价格字段来源`,
                               qd.created_at AS `创建时间`,
                               qd.updated_at AS `更新时间`
                        {base_from}
                        ORDER BY qd.quote_date DESC, qd.factory_id, qd.category_name, qd.id DESC
                        LIMIT %s OFFSET %s
                        """,
                        tuple(params) + (page_size, offset),
                    )
                    cols = [d[0] for d in cur.description]
                    rows = []
                    for r in cur.fetchall():
                        row: Dict[str, Any] = {}
                        for c, v in zip(cols, r):
                            if c == "价格字段来源":
                                row[c] = _json_cell_to_dict(v)
                            else:
                                row[c] = _cell_json(v)
                        rows.append(row)
            if response_format == "table":
                rows = [
                    {
                        "id": item["id"],
                        "日期": item["报价日期"],
                        "冶炼厂": item["冶炼厂"],
                        "品种": item["品类名"],
                        "基准价": item["普通价"],
                        "3%含税价": item["价格_3pct增值税"],
                        "13%含税价": item["价格_13pct增值税"],
                    }
                    for item in rows
                ]
            return {"code": 200, "data": {"total": total, "list": rows}}
        except ValueError:
            raise
        except Exception as e:
            logger.error(f"获取报价数据列表失败: {e}")
            raise

    # ==================== 接口5d：报价数据导出 Excel ====================

    def export_quote_details_excel(
        self,
        factory_id: Optional[int] = None,
        quote_date: Optional[str] = None,
        date_from: Optional[str] = None,
        date_to: Optional[str] = None,
        category_name: Optional[str] = None,
        category_exact: bool = False,
        max_rows: int = 50000,
    ) -> bytes:
        """与列表接口相同筛选条件，导出与表格列一致的 xlsx（最多 max_rows 行）。"""
        max_rows = min(max(max_rows, 1), 100000)
        where_sql, params = self._prepare_quote_details_filter(
            factory_id=factory_id,
            quote_date=quote_date,
            date_from=date_from,
            date_to=date_to,
            category_name=category_name,
            category_exact=category_exact,
        )
        base_from = (
            "FROM quote_details qd "
            "JOIN dict_factories df ON qd.factory_id = df.id "
            f"WHERE {where_sql}"
        )
        try:
            from openpyxl import Workbook

            with get_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        f"""
                        SELECT qd.quote_date,
                               df.name,
                               qd.category_name,
                               qd.unit_price,
                               qd.price_3pct_vat,
                               qd.price_13pct_vat
                        {base_from}
                        ORDER BY qd.quote_date DESC, qd.factory_id, qd.category_name, qd.id DESC
                        LIMIT %s
                        """,
                        tuple(params) + (max_rows,),
                    )
                    db_rows = cur.fetchall()

            wb = Workbook()
            ws = wb.active
            ws.title = "报价数据"
            ws.append(["日期", "冶炼厂", "品种", "基准价", "3%含税价", "13%含税价"])
            for row in db_rows:
                qd_d, fname, cname, up, p3, p13 = row
                ws.append(
                    [
                        qd_d.isoformat() if isinstance(qd_d, date) else qd_d,
                        fname,
                        cname,
                        _cell_json(up),
                        _cell_json(p3),
                        _cell_json(p13),
                    ]
                )
            buf = io.BytesIO()
            wb.save(buf)
            return buf.getvalue()
        except ValueError:
            raise
        except Exception as e:
            logger.error(f"导出报价 Excel 失败: {e}")
            raise

    # ==================== 接口7a：获取品类映射表 ====================

    def get_category_mapping(self) -> List[Dict[str, Any]]:
        try:
            with get_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        "SELECT row_id, category_id, name, is_main "
                        "FROM dict_categories "
                        "WHERE is_active = 1 "
                        "ORDER BY category_id, is_main DESC, row_id"
                    )
                    rows = cur.fetchall()

            result: Dict[int, Dict[str, Any]] = {}
            for row_id, cat_id, name, is_main in rows:
                if cat_id not in result:
                    result[cat_id] = {
                        "品类id": cat_id,
                        "品类名称": [],
                        "别名行": [],
                    }
                result[cat_id]["别名行"].append(
                    {
                        "行id": row_id,
                        "名称": name,
                        "是否主名称": bool(is_main),
                    }
                )
                if is_main:
                    result[cat_id]["品类名称"].insert(0, name)
                else:
                    result[cat_id]["品类名称"].append(name)

            return list(result.values())
        except Exception as e:
            logger.error(f"获取品类映射表失败: {e}")
            raise

    # ==================== 接口7：更新品类映射表 ====================

    @staticmethod
    def _normalize_category_mapping_names(names: List[str]) -> List[str]:
        norm: List[str] = []
        seen: set = set()
        for raw in names:
            if raw is None:
                continue
            n = str(raw).strip()
            if not n:
                raise ValueError("品类名称列表中含空名称")
            if len(n) > 50:
                raise ValueError(f"品种名长度不能超过 50: {n!r}")
            if n not in seen:
                seen.add(n)
                norm.append(n)
        if not norm:
            raise ValueError("品类名称列表不能为空")
        return norm

    @staticmethod
    def _resolve_replace_batch_name_owners(
        replace_rows: List[Tuple[int, int, List[str]]],
    ) -> Dict[str, int]:
        """
        同一次批量提交里，若同一品种名出现在多个「整组替换」条目中，全局只能落在一个 category_id
        （dict_categories.name UNIQUE）。优先保留在「本条名称数量更多」的分组；条数相同则保留在
        请求中更靠前的条目。返回值：name -> 归属 id；新品类（请求里 品类id<=0）用占位 id -(batch_idx+1)。
        """
        best: Dict[str, Tuple[int, int, int]] = {}
        for batch_idx, cid, norm in replace_rows:
            eff_cid = cid if cid > 0 else -(batch_idx + 1)
            glen = len(norm)
            for n in norm:
                if n not in best:
                    best[n] = (eff_cid, glen, batch_idx)
                else:
                    oc, og, ob = best[n]
                    if glen > og or (glen == og and batch_idx < ob):
                        best[n] = (eff_cid, glen, batch_idx)
        return {n: t[0] for n, t in best.items()}

    def update_category_mapping_batch(
        self,
        items: List[Tuple[int, List[str], bool]],
    ) -> Dict[str, Any]:
        """
        批量更新品类映射：先消解「一名多组」冲突，再逐条写入。
        items: (品类id, 品类名称列表, 仅追加别名)
        """
        normalized: List[Tuple[int, int, List[str], bool]] = []
        for batch_idx, (category_id, names, append_only) in enumerate(items):
            norm = self._normalize_category_mapping_names(names)
            normalized.append((batch_idx, category_id, norm, append_only))

        replace_rows: List[Tuple[int, int, List[str]]] = [
            (bi, cid, norm) for bi, cid, norm, app in normalized if not app
        ]
        owner_by_name = self._resolve_replace_batch_name_owners(replace_rows)

        last_cid: Optional[int] = None
        for batch_idx, category_id, norm, append_only in normalized:
            if append_only:
                r = self.update_category_mapping(
                    category_id=category_id,
                    names=norm,
                    append_only=True,
                )
                last_cid = r.get("品类id")
                continue

            eff_cid = category_id if category_id > 0 else -(batch_idx + 1)
            filtered = [n for n in norm if owner_by_name.get(n, eff_cid) == eff_cid]
            if not filtered:
                if category_id > 0:
                    try:
                        self.delete_category(category_id)
                    except ValueError:
                        pass
                continue
            norm = filtered

            r = self.update_category_mapping(
                category_id=category_id,
                names=norm,
                append_only=False,
            )
            last_cid = r.get("品类id")

        out: Dict[str, Any] = {
            "code": 200,
            "msg": "品类映射表更新成功，数据已存入数据库",
        }
        if last_cid is not None:
            out["品类id"] = last_cid
        return out

    def update_category_mapping(
        self,
        category_id: int,
        names: List[str],
        append_only: bool = False,
    ) -> Dict[str, Any]:
        norm = self._normalize_category_mapping_names(names)

        if append_only and category_id <= 0:
            raise ValueError("仅追加别名时 品类id 须为已有分组（>0）")

        had_active_before = False

        try:
            with get_conn() as conn:
                with conn.cursor() as cur:
                    if category_id <= 0:
                        cur.execute(
                            "SELECT COALESCE(MAX(category_id), 0) + 1 FROM dict_categories"
                        )
                        category_id = int(cur.fetchone()[0])
                    elif append_only:
                        cur.execute(
                            "SELECT name FROM dict_categories "
                            "WHERE category_id = %s AND is_active = 1 "
                            "ORDER BY is_main DESC, row_id ASC",
                            (category_id,),
                        )
                        existing_order = [row[0] for row in cur.fetchall()]
                        had_active_before = len(existing_order) > 0
                        merged: List[str] = []
                        seen_m: set = set()
                        for n in existing_order + norm:
                            if n not in seen_m:
                                seen_m.add(n)
                                merged.append(n)
                        norm = merged
                    else:
                        # 整组替换：该分组下原启用、且不在本次提交列表中的别名一律软删除
                        ph = ",".join(["%s"] * len(norm))
                        cur.execute(
                            f"UPDATE dict_categories SET is_active = 0 "
                            f"WHERE category_id = %s AND is_active = 1 "
                            f"AND name NOT IN ({ph})",
                            (category_id,) + tuple(norm),
                        )

                    # 将该 category_id 下所有旧记录的 is_main 置为 0
                    cur.execute(
                        "UPDATE dict_categories SET is_main = 0 WHERE category_id = %s",
                        (category_id,),
                    )

                    for i, name in enumerate(norm):
                        is_main = 1 if i == 0 else 0

                        cur.execute(
                            "SELECT row_id, category_id FROM dict_categories WHERE name = %s",
                            (name,),
                        )
                        existing = cur.fetchone()

                        if existing:
                            cur.execute(
                                "UPDATE dict_categories "
                                "SET category_id = %s, is_main = %s, is_active = 1 "
                                "WHERE row_id = %s",
                                (category_id, is_main, existing[0]),
                            )
                        else:
                            # 仅追加且分组原先已有启用别名：新插入的一律为别名，不得成为主名称
                            insert_main = (
                                0
                                if (append_only and had_active_before)
                                else is_main
                            )
                            cur.execute(
                                "INSERT INTO dict_categories "
                                "(category_id, name, is_main, is_active) "
                                "VALUES (%s, %s, %s, 1)",
                                (category_id, name, insert_main),
                            )

            return {
                "code": 200,
                "msg": "品类映射表更新成功，数据已存入数据库",
                "品类id": category_id,
            }

        except ValueError:
            raise
        except Exception as e:
            logger.error(f"更新品类映射失败: {e}")
            raise

    # ==================== 接口7b：按行修改品类别名 ====================

    def update_category_row(
        self,
        row_id: int,
        new_name: Optional[str] = None,
        set_main: Optional[bool] = None,
    ) -> Dict[str, Any]:
        if new_name is None and set_main is None:
            raise ValueError("至少需要提供 品种名 或 设为主名称（true）之一")

        try:
            with get_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        "SELECT row_id, category_id, name FROM dict_categories "
                        "WHERE row_id = %s AND is_active = 1",
                        (row_id,),
                    )
                    found = cur.fetchone()
                    if not found:
                        raise ValueError(f"品类别名不存在或已删除: 行id={row_id}")
                    _rid, cat_id, old_name = found

                    if new_name is not None:
                        nn = str(new_name).strip()
                        if not nn:
                            raise ValueError("品种名不能为空")
                        if len(nn) > 50:
                            raise ValueError("品种名长度不能超过 50")
                        cur.execute(
                            "SELECT row_id FROM dict_categories "
                            "WHERE name = %s AND row_id <> %s AND is_active = 1",
                            (nn, row_id),
                        )
                        if cur.fetchone():
                            raise ValueError(f"品种名「{nn}」已被其它别名使用")
                        cur.execute(
                            "UPDATE dict_categories SET name = %s WHERE row_id = %s",
                            (nn, row_id),
                        )
                        cur.execute(
                            "UPDATE quote_details SET category_name = %s WHERE category_name = %s",
                            (nn, old_name),
                        )

                    if set_main is True:
                        cur.execute(
                            "UPDATE dict_categories SET is_main = 0 WHERE category_id = %s",
                            (cat_id,),
                        )
                        cur.execute(
                            "UPDATE dict_categories SET is_main = 1 WHERE row_id = %s",
                            (row_id,),
                        )

            return {"code": 200, "msg": "品类别名已更新"}
        except ValueError:
            raise
        except Exception as e:
            logger.error(f"修改品类别名失败: {e}")
            raise

    # ==================== 接口7c：删除品类分组（软删除） ====================

    def delete_category(self, category_id: int) -> Dict[str, Any]:
        if category_id < 1:
            raise ValueError("品类id 无效")
        try:
            with get_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        "UPDATE dict_categories SET is_active = 0 "
                        "WHERE category_id = %s AND is_active = 1",
                        (category_id,),
                    )
                    n = cur.rowcount
                    if n == 0:
                        raise ValueError(f"品类 id={category_id} 不存在或已删除")
            return {"code": 200, "msg": "品类分组已删除", "影响行数": n}
        except ValueError:
            raise
        except Exception as e:
            logger.error(f"删除品类分组失败: {e}")
            raise

    # ==================== 接口7d：删除单条品类别名（软删除） ====================

    def delete_category_row(self, row_id: int) -> Dict[str, Any]:
        if row_id < 1:
            raise ValueError("行id 无效")
        try:
            with get_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        "SELECT category_id, is_main FROM dict_categories "
                        "WHERE row_id = %s AND is_active = 1",
                        (row_id,),
                    )
                    row = cur.fetchone()
                    if not row:
                        raise ValueError(f"品类别名不存在或已删除: 行id={row_id}")
                    cat_id, was_main = int(row[0]), int(row[1])

                    cur.execute(
                        "UPDATE dict_categories SET is_active = 0 WHERE row_id = %s",
                        (row_id,),
                    )

                    if was_main:
                        cur.execute(
                            "SELECT row_id FROM dict_categories "
                            "WHERE category_id = %s AND is_active = 1 "
                            "ORDER BY row_id ASC LIMIT 1",
                            (cat_id,),
                        )
                        nxt = cur.fetchone()
                        if nxt:
                            cur.execute(
                                "UPDATE dict_categories SET is_main = 1 WHERE row_id = %s",
                                (nxt[0],),
                            )

            return {"code": 200, "msg": "品类别名已删除"}
        except ValueError:
            raise
        except Exception as e:
            logger.error(f"删除品类别名失败: {e}")
            raise

    # ==================== 接口A7：采购建议 ====================

    def get_purchase_suggestion(
        self,
        warehouse_ids: List[int],
        demands: List[Dict[str, Any]],
        price_type: Optional[str] = None,
    ) -> Dict[str, Any]:
        """
        根据仓库列表和需求（品类+吨数），查询最新运费和报价。
        冶炼厂默认取 dict_factories 中全部启用冶炼厂，无需前端传入。
        整理结构化数据后调用 LLM 生成各仓库发车建议表。
        price_type: 目标税率类型，None=普通价, 1pct/3pct/13pct/normal_invoice/reverse_invoice
        """
        if not warehouse_ids or not demands:
            raise ValueError("仓库列表和需求不能为空")

        # price_type → (quote_details列名, 展示名)
        PRICE_COL_MAP = {
            None:             ("unit_price",            "普通价"),
            "1pct":           ("price_1pct_vat",        "1%增值税"),
            "3pct":           ("price_3pct_vat",        "3%增值税"),
            "13pct":          ("price_13pct_vat",       "13%增值税"),
            "normal_invoice": ("price_normal_invoice",  "普通发票"),
            "reverse_invoice":("price_reverse_invoice", "反向发票"),
        }
        VAT_TAX_TYPE_MAP = {"1pct": "1pct", "3pct": "3pct", "13pct": "13pct"}

        if price_type not in PRICE_COL_MAP:
            raise ValueError(f"不支持的 price_type: {price_type}")

        target_col, price_type_name = PRICE_COL_MAP[price_type]
        target_tax = VAT_TAX_TYPE_MAP.get(price_type)

        category_ids = list({d["category_id"] for d in demands})

        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT id FROM dict_factories WHERE is_active = 1 ORDER BY id"
                )
                smelter_ids = [r[0] for r in cur.fetchall()]
                if not smelter_ids:
                    raise ValueError("没有可用的冶炼厂，请先在 dict_factories 中维护启用冶炼厂")

                wh_ph = ",".join(["%s"] * len(warehouse_ids))
                sm_ph = ",".join(["%s"] * len(smelter_ids))
                cat_ph = ",".join(["%s"] * len(category_ids))

                # 仓库名称
                cur.execute(
                    f"SELECT id, name FROM dict_warehouses WHERE id IN ({wh_ph})",
                    tuple(warehouse_ids),
                )
                warehouse_name_map: Dict[int, str] = {r[0]: r[1] for r in cur.fetchall()}

                # 品类主名称
                cur.execute(
                    f"SELECT category_id, "
                    f"COALESCE(MAX(CASE WHEN is_main=1 THEN name END), MAX(name)) "
                    f"FROM dict_categories "
                    f"WHERE category_id IN ({cat_ph}) AND is_active=1 "
                    f"GROUP BY category_id",
                    tuple(category_ids),
                )
                cat_name_map: Dict[int, str] = {r[0]: r[1] for r in cur.fetchall()}

                # 冶炼厂名称
                cur.execute(
                    f"SELECT id, name FROM dict_factories WHERE id IN ({sm_ph})",
                    tuple(smelter_ids),
                )
                factory_name_map: Dict[int, str] = {r[0]: r[1] for r in cur.fetchall()}

                # 最新运费：每个(仓库, 冶炼厂)取最新日期，保留仓库维度
                cur.execute(
                    f"""
                    SELECT dw.id AS wid, dw.name AS wname,
                           df.id AS fid, df.name AS fname,
                           fr.price_per_ton
                    FROM freight_rates fr
                    JOIN dict_warehouses dw ON fr.warehouse_id = dw.id
                    JOIN dict_factories  df ON fr.factory_id  = df.id
                    WHERE dw.id IN ({wh_ph})
                      AND df.id IN ({sm_ph})
                      AND fr.effective_date = (
                          SELECT MAX(fr2.effective_date)
                          FROM freight_rates fr2
                          WHERE fr2.factory_id  = fr.factory_id
                            AND fr2.warehouse_id = fr.warehouse_id
                      )
                    """,
                    tuple(warehouse_ids) + tuple(smelter_ids),
                )
                # freight_map: {(warehouse_id, factory_id): freight}
                freight_map: Dict[tuple, float] = {
                    (r[0], r[2]): (float(r[4]) if r[4] is not None else 0.0) for r in cur.fetchall()
                }

                # 税率表
                cur.execute(
                    f"SELECT factory_id, tax_type, tax_rate "
                    f"FROM factory_tax_rates WHERE factory_id IN ({sm_ph})",
                    tuple(smelter_ids),
                )
                tax_rate_map: Dict[int, Dict[str, float]] = {}
                for fid, ttype, rate in cur.fetchall():
                    tax_rate_map.setdefault(fid, {})[ttype] = float(rate)

                # category_id → 品类名称列表
                cur.execute(
                    f"SELECT category_id, name FROM dict_categories "
                    f"WHERE category_id IN ({cat_ph}) AND is_active = 1",
                    tuple(category_ids),
                )
                cat_id_to_names: Dict[int, List[str]] = {}
                for cat_id, name in cur.fetchall():
                    cat_id_to_names.setdefault(cat_id, []).append(name)

                if not cat_id_to_names:
                    return {"demand_rows": [], "raw": []}

                # 所有品类名称
                all_cat_names = [name for names in cat_id_to_names.values() for name in names]
                cn_ph = ",".join(["%s"] * len(all_cat_names))

                # 最新报价：通过品类名称查询
                cur.execute(
                    f"""
                    SELECT factory_id, category_name,
                           unit_price, price_1pct_vat, price_3pct_vat, price_13pct_vat,
                           price_normal_invoice, price_reverse_invoice
                    FROM quote_details
                    WHERE factory_id IN ({sm_ph})
                      AND category_name IN ({cn_ph})
                      AND quote_date = (
                          SELECT MAX(qd2.quote_date)
                          FROM quote_details qd2
                          WHERE qd2.factory_id = quote_details.factory_id
                            AND qd2.category_name = quote_details.category_name
                      )
                    """,
                    tuple(smelter_ids) + tuple(all_cat_names),
                )
                col_names = ["unit_price", "price_1pct_vat", "price_3pct_vat",
                             "price_13pct_vat", "price_normal_invoice", "price_reverse_invoice"]
                raw_price_map: Dict[tuple, Dict[str, Optional[float]]] = {}
                for row in cur.fetchall():
                    fid_r, cat_name = row[0], row[1]
                    raw_price_map[(fid_r, cat_name)] = {
                        col: (float(v) if v is not None else None)
                        for col, v in zip(col_names, row[2:])
                    }

        # 价格反算逻辑
        COL_TO_TAX: Dict[str, str] = {
            "price_1pct_vat": "1pct",
            "price_3pct_vat": "3pct",
            "price_13pct_vat": "13pct",
        }

        def resolve_price(fid: int, cat_id: int) -> Optional[float]:
            cat_names = cat_id_to_names.get(cat_id, [])
            for cat_name in cat_names:
                prices = raw_price_map.get((fid, cat_name), {})
                if not prices:
                    continue

                rates = tax_rate_map.get(fid, {})
                merged = merge_factory_rates(rates)

                direct = prices.get(target_col)
                if direct is not None:
                    return direct

                if target_tax and prices.get("unit_price") is not None and target_tax in merged:
                    return inclusive_from_net(float(prices["unit_price"]), merged[target_tax])

                if target_col == "unit_price":
                    for col, src_tax in COL_TO_TAX.items():
                        known_price = prices.get(col)
                        if known_price is not None and src_tax in merged:
                            net = net_from_inclusive(float(known_price), merged[src_tax])
                            return round(net, 2)

                if target_tax and target_tax in merged:
                    for col, src_tax in COL_TO_TAX.items():
                        known_price = prices.get(col)
                        if known_price is not None and src_tax in merged:
                            net = net_from_inclusive(float(known_price), merged[src_tax])
                            return inclusive_from_net(net, merged[target_tax])

            return None

        # 构建 price_map: {(factory_id, category_id): price}
        price_map: Dict[tuple, Optional[float]] = {}
        for fid in smelter_ids:
            for cid in category_ids:
                price_map[(fid, cid)] = resolve_price(fid, cid)

        # 构造结构化数据：每条需求 × 全部冶炼厂，报价与各仓库运费对比
        # 与 get_comparison 一致：比价利润 = 报价×吨数 − 运费×吨数 → 元/吨档为 (报价 − 运费)
        demand_rows = []
        raw = []
        for d in demands:
            cid = d["category_id"]
            demand_tons = float(d["demand"])
            for fid in smelter_ids:
                fname = factory_name_map.get(fid, f"冶炼厂{fid}")
                cat_name = cat_name_map.get(cid, f"品类{cid}")
                price = price_map.get((fid, cid))

                warehouse_options = []
                for wid in warehouse_ids:
                    wname = warehouse_name_map.get(wid, f"仓库{wid}")
                    freight = freight_map.get((wid, fid))
                    margin_per_ton: Optional[float] = None
                    if price is not None and freight is not None:
                        margin_per_ton = round(float(price) - float(freight), 2)
                    profit_yuan: Optional[float] = None
                    if margin_per_ton is not None:
                        profit_yuan = round(margin_per_ton * demand_tons, 2)
                    warehouse_options.append({
                        "仓库": wname,
                        "运费(元/吨)": freight,
                        "比价利润元每吨": margin_per_ton,
                        "比价利润(元)": profit_yuan,
                    })
                    raw.append({
                        "冶炼厂": fname,
                        "品类": cat_name,
                        "需求吨数": demand_tons,
                        "报价(元/吨)": price,
                        "仓库": wname,
                        "运费(元/吨)": freight,
                        "比价利润元每吨": margin_per_ton,
                        "比价利润(元)": profit_yuan,
                    })

                demand_rows.append({
                    "冶炼厂": fname,
                    "品类": cat_name,
                    "需求吨数(吨)": demand_tons,
                    "报价(元/吨)": price,
                    "各仓库运费对比": warehouse_options,
                })

        # 构造 prompt，调用大模型（OpenAI 兼容协议）
        import json
        from openai import OpenAI
        from app import config as app_config

        if not (app_config.LLM_API_KEY or "").strip():
            raise ValueError(
                "未配置文本大模型密钥，无法生成采购建议。请设置 LLM_API_KEY；若与报价图识别共用阿里云百炼，"
                "也可只配 VLM_API_KEY（或 DASHSCOPE_API_KEY / QWEN_API_KEY），"
                "此时默认使用百炼兼容端点与 qwen-plus。其它厂商请显式配置 LLM_API_KEY、LLM_BASE_URL、LLM_MODEL。"
            )

        client = OpenAI(api_key=app_config.LLM_API_KEY, base_url=app_config.LLM_BASE_URL)
        data_str = json.dumps(demand_rows, ensure_ascii=False, indent=2)
        prompt = f"""以下是各需求的报价及各仓库运费数据：

{data_str}

请给出各仓库发车建议，要求：
1. 与系统比价一致：每条线路的「比价利润(元)」= 报价×吨数 − 运费×吨数（数据中已按此计算）；优先选比价利润更高（更优）的仓库
2. 同仓库不同品类可混装，尽量整车（20-30吨）
3. 按仓库分段输出：仓库名、装车方案（品类+吨数+冶炼厂+比价利润）、备注
4. 数据缺失的在备注注明
5. 纯文本，简洁"""

        try:
            resp = client.chat.completions.create(
                model=app_config.LLM_MODEL,
                max_tokens=2048,
                messages=[{"role": "user", "content": prompt}],
            )
            suggestion = resp.choices[0].message.content
        except Exception as exc:
            err_text = str(exc).lower()
            status = getattr(exc, "status_code", None)
            logger.exception("采购建议大模型调用失败")
            if status == 403 or ("403" in str(exc) and "forbidden" in err_text):
                raise PurchaseSuggestionLLMError(
                    "大模型服务端拒绝请求（HTTP 403）。常见原因：API Key 无效或无权访问该模型、"
                    "LLM_BASE_URL 与密钥不属于同一服务商、控制台 IP 白名单未放行当前服务器、"
                    "套餐/配额或地域策略限制。请在部署环境检查 LLM_API_KEY、LLM_BASE_URL、LLM_MODEL，"
                    "并登录模型服务商控制台核对权限与网络策略。"
                ) from exc
            raise PurchaseSuggestionLLMError(
                f"大模型调用失败，无法生成建议正文。原始错误：{exc}"
            ) from exc

        return {"code": 200, "data": {"suggestion": suggestion, "raw": raw}}


# ==================== 单例工厂 ====================

_tl_service: Optional[TLService] = None


def get_tl_service() -> TLService:
    global _tl_service
    if _tl_service is None:
        _tl_service = TLService()
    return _tl_service
