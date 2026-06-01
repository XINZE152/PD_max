"""库房库存/收货价格 Excel 解析与导入匹配逻辑单元测试。"""
import io
import sys
import unittest
from decimal import Decimal
from unittest.mock import MagicMock

from openpyxl import Workbook

_vlm_stub = MagicMock()
_vlm_stub.QwenVLFullExtractor = MagicMock
_vlm_stub.VLMConfig = MagicMock
sys.modules.setdefault("app.services.vlm_extractor_service", _vlm_stub)

from app.services.tl_service import (  # noqa: E402
    TLService,
    _aggregate_import_skip_reasons,
    _rank_warehouse_name_match,
)
from app.services.warehouse_inventory_excel import (  # noqa: E402
    WarehouseInventoryExcelError,
    parse_warehouse_inventory_workbook,
)
from app.services.warehouse_receipt_price_excel import (  # noqa: E402
    parse_warehouse_receipt_price_workbook,
)


class _CategoryCursor:
    """模拟 dict_categories 查询，供品类解析测试。"""

    def __init__(self, active_rows: list[tuple[int, str]]) -> None:
        self._active = active_rows
        self._last_sql = ""
        self._last_params: tuple = ()

    def execute(self, sql: str, params=None) -> None:
        self._last_sql = sql
        self._last_params = tuple(params or ())

    def fetchone(self):
        if "name = %s AND is_active = 1" in self._last_sql:
            name = self._last_params[0]
            for cid, n in self._active:
                if n == name:
                    return (cid,)
            return None
        if "is_active FROM dict_categories WHERE name = %s" in self._last_sql:
            return None
        return None

    def fetchall(self):
        if "SELECT category_id, name FROM dict_categories WHERE is_active = 1" in self._last_sql:
            return list(self._active)
        return []


class WarehouseInventoryExcelTests(unittest.TestCase):
    def test_parse_inventory_workbook(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "导入数据"
        ws.append(["库房名称", "回收品种", "当前库存", "库存日期"])
        ws.append(["测试库房", "电动电瓶", 88.5, "2026-05-29"])
        buf = io.BytesIO()
        wb.save(buf)
        rows, meta = parse_warehouse_inventory_workbook(buf.getvalue())
        self.assertEqual(meta["parsed_rows"], 1)
        self.assertEqual(rows[0].warehouse_name, "测试库房")
        self.assertEqual(rows[0].category_name, "电动电瓶")
        self.assertEqual(rows[0].inventory_ton, Decimal("88.5"))

    def test_parse_inventory_requires_category_column(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "导入数据"
        ws.append(["库房名称", "当前库存", "库存日期"])
        ws.append(["测试库房", 88.5, "2026-05-29"])
        buf = io.BytesIO()
        wb.save(buf)
        with self.assertRaises(Exception):
            parse_warehouse_inventory_workbook(buf.getvalue())

    def test_parse_inventory_rejects_receipt_price_header(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "导入数据"
        ws.append(["库房名称", "回收品种", "价格"])
        ws.append(["测试库房", "电轿电瓶", 9400])
        buf = io.BytesIO()
        wb.save(buf)
        with self.assertRaises(WarehouseInventoryExcelError) as ctx:
            parse_warehouse_inventory_workbook(buf.getvalue())
        self.assertIn("收货价格格式", str(ctx.exception))
        self.assertIn("当前库存", str(ctx.exception))


class WarehouseReceiptPriceExcelTests(unittest.TestCase):
    def test_parse_receipt_price_workbook(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "导入数据"
        ws.append(["库房名称", "回收品种", "价格", "价格日期"])
        ws.append(["测试库房", "电动电瓶", 15200, "2026-06-01"])
        buf = io.BytesIO()
        wb.save(buf)
        rows, meta = parse_warehouse_receipt_price_workbook(buf.getvalue())
        self.assertEqual(meta["parsed_rows"], 1)
        self.assertEqual(rows[0].warehouse_name, "测试库房")
        self.assertEqual(rows[0].category_name, "电动电瓶")
        self.assertEqual(rows[0].price_per_ton, Decimal("15200"))
        self.assertEqual(rows[0].price_date.isoformat(), "2026-06-01")


class WarehouseImportMatchTests(unittest.TestCase):
    def setUp(self) -> None:
        self.service = TLService()

    def test_resolve_dian_dong_dian_ping_via_difflib(self) -> None:
        cur = _CategoryCursor(
            [
                (6, "电动车电瓶"),
                (6, "电瓶车电瓶"),
                (12, "电轿电瓶"),
            ]
        )
        cat_id = self.service._resolve_category_id_by_name(cur, "电动电瓶")
        self.assertEqual(cat_id, 6)

    def test_resolve_exact_category(self) -> None:
        cur = _CategoryCursor([(12, "电轿电瓶")])
        self.assertEqual(
            self.service._resolve_category_id_by_name(cur, "电轿电瓶"), 12
        )

    def test_fuzzy_warehouse_substring_match(self) -> None:
        candidates = [
            (100, "荥阳世纪再生资源有限公司"),
            (200, "河北发江废旧物资回收有限公司"),
        ]
        wid, kind = TLService._fuzzy_match_warehouse_id_from_candidates(
            "荥阳世纪再生资源", candidates
        )
        self.assertEqual(wid, 100)
        self.assertEqual(kind, "fuzzy")

    def test_rank_warehouse_exact_before_substring(self) -> None:
        self.assertLess(
            _rank_warehouse_name_match("测试库", "测试库")[0],
            _rank_warehouse_name_match("测试库", "测试库房有限公司")[0],
        )

    def test_aggregate_skip_reasons(self) -> None:
        errors = [
            "第 2 行：回收品种「电动电瓶」不存在",
            "第 3 行：库房名称「未知库」不存在",
            "第 4 行：库房名称「重复库」匹配到多个库房",
        ]
        reasons = _aggregate_import_skip_reasons(errors)
        self.assertEqual(reasons.get("品类不存在"), 1)
        self.assertEqual(reasons.get("库房不存在"), 1)
        self.assertEqual(reasons.get("库房歧义"), 1)


if __name__ == "__main__":
    unittest.main()
